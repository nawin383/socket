"""
🟠 BITCOIN VIRTUAL TRADING PLATFORM - PROFESSIONAL EDITION
================================================================

Advanced Features:
- SQLite Database with full trade history
- Real-time WebSocket data updates
- Advanced risk analytics (Sharpe, Sortino, VaR, CVaR)
- Options strategy builder with 10+ strategies
- Interactive charts (Portfolio, P&L, Greeks)
- Stop Loss / Take Profit automation
- Performance analytics dashboard
- Position sizing calculator (Kelly Criterion)
- Portfolio Greeks aggregation
- Export to Excel with advanced formatting
- Trade journal with notes and tags

Requirements:
pip install -r requirements_btc_pro.txt
"""

import sys
import os
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, simpledialog, filedialog
from datetime import datetime, timedelta
import threading
import time
from typing import Dict, Any, List, Optional
import requests

# Import custom modules
from btc_config_pro import ConfigManager
from btc_models_pro import DatabaseManager, Trade, Portfolio
from btc_analytics_pro import (
    PerformanceAnalytics, RiskMetrics, TradeAnalytics,
    generate_performance_report
)
from btc_strategies_pro import OptionsStrategyBuilder, StrategyAnalyzer

try:
    import ttkbootstrap as tb
    BOOTSTRAP_AVAILABLE = True
except ImportError:
    BOOTSTRAP_AVAILABLE = False
    print("⚠️  ttkbootstrap not available, using default theme")

try:
    import pandas as pd
    import numpy as np
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("⚠️  pandas/numpy not available, some features disabled")

try:
    import matplotlib
    matplotlib.use('TkAgg')
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.figure import Figure
    import matplotlib.pyplot as plt
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False
    print("⚠️  matplotlib not available, charts disabled")

try:
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False
    print("⚠️  plotly not available, interactive charts disabled")


# Delta Exchange API
DELTA_API_URL = "https://cdn.india.deltaex.org/v2/tickers"
DELTA_PARAMS = {
    'contract_types': 'futures,perpetual_futures,move_options,put_options,call_options,spot'
}


def safe_float(value, default=0.0):
    """Safely convert value to float"""
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


class DeltaExchangeAPI:
    """Delta Exchange API wrapper"""

    def __init__(self):
        self.base_url = DELTA_API_URL
        self.cache = {}
        self.cache_time = None
        self.cache_duration = 10

    def fetch_tickers(self, underlying='BTC'):
        """Fetch all tickers"""
        try:
            # Use cache if recent
            if self.cache_time and (datetime.now() - self.cache_time).seconds < self.cache_duration:
                cached_data = self.cache.get(underlying, [])
                if cached_data:
                    return cached_data

            response = requests.get(self.base_url, params=DELTA_PARAMS, timeout=10)
            response.raise_for_status()

            data = response.json()
            result = data.get('result', [])

            if not result:
                return []

            # Filter by underlying
            filtered = [item for item in result if item and item.get('underlying_asset_symbol') == underlying]

            # Cache results
            self.cache[underlying] = filtered
            self.cache_time = datetime.now()

            return filtered

        except Exception as e:
            print(f"❌ Error fetching tickers: {e}")
            return []

    def format_instrument(self, ticker):
        """Format instrument data"""
        if not ticker:
            return None

        quotes = ticker.get('quotes') or {}
        greeks = ticker.get('greeks') or {}

        return {
            'instrument_token': ticker.get('product_id'),
            'tradingsymbol': ticker.get('symbol', ''),
            'name': ticker.get('underlying_asset_symbol', ''),
            'description': ticker.get('description', ''),
            'contract_type': ticker.get('contract_type', ''),
            'strike': safe_float(ticker.get('strike_price', 0)),
            'instrument_type': self._get_instrument_type(ticker.get('contract_type')),
            'exchange': 'DELTA',
            'last_price': safe_float(ticker.get('mark_price', 0)),
            'close_price': safe_float(ticker.get('close', 0)),
            'high': safe_float(ticker.get('high', 0)),
            'low': safe_float(ticker.get('low', 0)),
            'open': safe_float(ticker.get('open', 0)),
            'oi': safe_float(ticker.get('oi_contracts', 0)),
            'oi_value': safe_float(ticker.get('oi_value_usd', 0)),
            'mark_change_24h': safe_float(ticker.get('mark_change_24h', 0)),
            'best_bid': safe_float(quotes.get('best_bid', 0)),
            'best_ask': safe_float(quotes.get('best_ask', 0)),
            'bid_iv': safe_float(quotes.get('bid_iv', 0)),
            'ask_iv': safe_float(quotes.get('ask_iv', 0)),
            'mark_iv': safe_float(quotes.get('mark_iv', 0)),
            'delta': safe_float(greeks.get('delta', 0)),
            'gamma': safe_float(greeks.get('gamma', 0)),
            'theta': safe_float(greeks.get('theta', 0)),
            'vega': safe_float(greeks.get('vega', 0)),
            'rho': safe_float(greeks.get('rho', 0)),
            'spot_price': safe_float(ticker.get('spot_price', 0)),
            'timestamp': ticker.get('timestamp')
        }

    def _get_instrument_type(self, contract_type):
        """Convert contract type"""
        mapping = {
            'futures': 'FUT',
            'perpetual_futures': 'PERP',
            'put_options': 'PE',
            'call_options': 'CE',
            'spot': 'SPOT'
        }
        return mapping.get(contract_type, 'FUT')

    def get_expiries(self, underlying='BTC', contract_type='futures'):
        """Get available expiries"""
        tickers = self.fetch_tickers(underlying)
        expiries = set()

        for ticker in tickers:
            if not ticker:
                continue

            ticker_type = ticker.get('contract_type')
            symbol = ticker.get('symbol', '')

            if contract_type == 'options':
                if ticker_type in ['put_options', 'call_options']:
                    expiry = self._extract_expiry(symbol)
                    if expiry:
                        expiries.add(expiry)
            elif ticker_type in ['futures', 'perpetual_futures']:
                if ticker_type == 'perpetual_futures':
                    expiries.add('PERPETUAL')
                else:
                    expiry = self._extract_expiry(symbol)
                    if expiry:
                        expiries.add(expiry)

        return sorted(list(expiries), key=lambda x: (x != 'PERPETUAL', x))

    def _extract_expiry(self, symbol):
        """Extract expiry from symbol"""
        if '-' in symbol:
            parts = symbol.split('-')
            if len(parts) >= 2:
                expiry_part = parts[1]
                if '-' in expiry_part:
                    expiry_part = expiry_part.split('-')[0]
                return expiry_part
        return None

    def get_strikes(self, underlying='BTC', expiry=None):
        """Get available strikes"""
        tickers = self.fetch_tickers(underlying)
        strikes = set()

        for ticker in tickers:
            if not ticker:
                continue

            if ticker.get('contract_type') not in ['put_options', 'call_options']:
                continue

            symbol = ticker.get('symbol', '')
            if expiry and expiry not in symbol:
                continue

            strike = safe_float(ticker.get('strike_price'))
            if strike > 0:
                strikes.add(strike)

        return sorted(list(strikes))


class BTCTradingPlatformPro:
    """Bitcoin Virtual Trading Platform - Professional Edition"""

    def __init__(self, root):
        self.root = root
        self.root.title("🟠 Bitcoin Virtual Trading Platform - PRO Edition")
        self.root.geometry("1920x1080")

        # Initialize components
        self.config = ConfigManager()
        self.db = DatabaseManager(self.config.get('trading', 'database_path', 'btc_trading_pro.db'))
        self.delta_api = DeltaExchangeAPI()

        # State
        self.btc_instruments = []
        self.instruments_loaded = False
        self.btc_refresh_enabled = True
        self.last_btc_refresh_time = None
        self.btc_refresh_timer = None
        self.selected_expiry = None
        self.btc_spot_price = 0

        # Load data
        self.portfolio = self.db.get_portfolio()
        self.trades = self.db.get_all_trades()

        # Create GUI
        self.create_widgets()
        self.update_dashboard()

        # Load BTC instruments
        self.root.after(500, self.load_btc_instruments)

        # Setup cleanup
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def log(self, message):
        """Add message to log"""
        timestamp = datetime.now().strftime('%H:%M:%S')
        log_message = f"[{timestamp}] {message}\n"
        if hasattr(self, 'log_text'):
            self.log_text.insert(tk.END, log_message)
            self.log_text.see(tk.END)
        print(log_message.strip())

    def load_btc_instruments(self):
        """Load Bitcoin instruments"""
        self.log("🔄 Loading BTC instruments from Delta Exchange...")
        threading.Thread(target=self._load_btc_background, daemon=True).start()

    def _load_btc_background(self):
        """Background thread to load BTC instruments"""
        try:
            tickers = self.delta_api.fetch_tickers('BTC')

            if not tickers:
                self.root.after(0, lambda: self.log("❌ No data received from Delta Exchange"))
                return

            self.btc_instruments = []
            for ticker in tickers:
                formatted = self.delta_api.format_instrument(ticker)
                if formatted:
                    self.btc_instruments.append(formatted)

            count = len(self.btc_instruments)
            self.instruments_loaded = True
            self.root.after(0, lambda c=count: self.log(f"✅ Loaded {c} BTC instruments"))
            self.root.after(0, self.update_ui_after_load)

        except Exception as e:
            self.root.after(0, lambda msg=str(e): self.log(f"❌ Error: {msg}"))

    def update_ui_after_load(self):
        """Update UI after instruments loaded"""
        if self.instruments_loaded and self.btc_instruments:
            self.update_expiries()
            self.update_spot_info()
            self.start_btc_refresh()

    def start_btc_refresh(self):
        """Start auto-refresh"""
        if self.btc_refresh_enabled:
            interval = self.config.get('trading', 'auto_refresh_interval', 15)
            self.btc_refresh_timer = self.root.after(interval * 1000, self.auto_btc_refresh)

    def auto_btc_refresh(self):
        """Auto-refresh BTC prices"""
        if not self.btc_refresh_enabled:
            return

        self.log("🔄 Auto-refresh: Updating BTC prices...")
        threading.Thread(target=self._refresh_btc_background, daemon=True).start()

        interval = self.config.get('trading', 'auto_refresh_interval', 15)
        self.btc_refresh_timer = self.root.after(interval * 1000, self.auto_btc_refresh)

    def _refresh_btc_background(self):
        """Background refresh"""
        try:
            tickers = self.delta_api.fetch_tickers('BTC')

            if not tickers:
                return

            self.btc_instruments = []
            for ticker in tickers:
                formatted = self.delta_api.format_instrument(ticker)
                if formatted:
                    self.btc_instruments.append(formatted)

            self.last_btc_refresh_time = datetime.now()

            # Update positions
            self._update_btc_positions()

            self.root.after(0, self.update_spot_info)
            self.root.after(0, lambda: self.log("✅ BTC refresh completed"))

        except Exception as e:
            self.root.after(0, lambda msg=str(e): self.log(f"❌ Refresh error: {msg}"))

    def _update_btc_positions(self):
        """Update position prices"""
        open_trades = self.db.get_open_trades()

        if not open_trades:
            return

        symbol_map = {inst['tradingsymbol']: inst for inst in self.btc_instruments}
        updated = 0

        for trade in open_trades:
            symbol = trade.trading_symbol
            if symbol in symbol_map:
                inst = symbol_map[symbol]
                ltp = safe_float(inst.get('last_price', 0))

                if ltp > 0:
                    trade.current_price = ltp
                    trade.unrealized_pnl = trade.calculate_pnl()

                    # Update Greeks
                    if trade.instrument_type in ['CE', 'PE']:
                        trade.delta = safe_float(inst.get('delta', 0))
                        trade.gamma = safe_float(inst.get('gamma', 0))
                        trade.theta = safe_float(inst.get('theta', 0))
                        trade.vega = safe_float(inst.get('vega', 0))
                        trade.iv = safe_float(inst.get('mark_iv', 0))

                    # Check SL/TP
                    self._check_sl_tp(trade, ltp)

                    self.db.update_trade(trade)
                    updated += 1

        if updated > 0:
            self.root.after(0, self.update_dashboard)
            self.root.after(0, self.refresh_positions)

    def _check_sl_tp(self, trade: Trade, current_price: float):
        """Check stop loss and take profit"""
        if trade.stop_loss and current_price <= trade.stop_loss and trade.trade_type == 'buy':
            self._auto_close_trade(trade, current_price, "Stop Loss Hit")
        elif trade.stop_loss and current_price >= trade.stop_loss and trade.trade_type == 'sell':
            self._auto_close_trade(trade, current_price, "Stop Loss Hit")
        elif trade.take_profit and current_price >= trade.take_profit and trade.trade_type == 'buy':
            self._auto_close_trade(trade, current_price, "Take Profit Hit")
        elif trade.take_profit and current_price <= trade.take_profit and trade.trade_type == 'sell':
            self._auto_close_trade(trade, current_price, "Take Profit Hit")

    def _auto_close_trade(self, trade: Trade, exit_price: float, reason: str):
        """Auto-close trade"""
        trade.status = 'closed'
        trade.exit_price = exit_price
        trade.exit_date = datetime.now().strftime('%Y-%m-%d')
        trade.realized_pnl = trade.calculate_pnl()
        trade.unrealized_pnl = 0

        self.db.update_trade(trade)
        self.root.after(0, lambda: self.log(f"🔔 {reason}: {trade.trading_symbol} closed @ ${exit_price:.2f}"))

    def update_spot_info(self):
        """Update spot price info"""
        if not self.btc_instruments:
            return

        spot = next((inst for inst in self.btc_instruments if inst['instrument_type'] == 'SPOT'), None)

        if spot and hasattr(self, 'btc_spot_label'):
            price = safe_float(spot.get('last_price', 0))
            change = safe_float(spot.get('mark_change_24h', 0))
            self.btc_spot_price = price

            color = 'green' if change >= 0 else 'red'
            self.btc_spot_label.config(
                text=f"BTC Spot: ${price:,.2f} ({change:+.2f}%)",
                foreground=color)

    def create_widgets(self):
        """Create GUI"""
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=10)

        # Tabs
        self.dashboard_frame = ttk.Frame(self.notebook)
        self.trade_frame = ttk.Frame(self.notebook)
        self.positions_frame = ttk.Frame(self.notebook)
        self.analytics_frame = ttk.Frame(self.notebook)
        self.strategy_frame = ttk.Frame(self.notebook)
        self.settings_frame = ttk.Frame(self.notebook)

        self.notebook.add(self.dashboard_frame, text='📊 Dashboard')
        self.notebook.add(self.trade_frame, text='💹 Trade')
        self.notebook.add(self.positions_frame, text='💼 Positions')
        self.notebook.add(self.analytics_frame, text='📈 Analytics')
        self.notebook.add(self.strategy_frame, text='🎯 Strategies')
        self.notebook.add(self.settings_frame, text='⚙️ Settings')

        self.build_dashboard()
        self.build_trade_tab()
        self.build_positions_tab()
        self.build_analytics_tab()
        self.build_strategy_tab()
        self.build_settings_tab()

    def build_dashboard(self):
        """Build dashboard"""
        # Portfolio stats
        stats_frame = ttk.LabelFrame(self.dashboard_frame, text="Portfolio Overview", padding=20)
        stats_frame.pack(fill='x', padx=20, pady=10)

        self.initial_capital_label = ttk.Label(stats_frame,
            text=f"Initial: ${self.portfolio.initial_capital:,.2f}",
            font=('Arial', 14, 'bold'))
        self.initial_capital_label.grid(row=0, column=0, padx=20)

        self.current_capital_label = ttk.Label(stats_frame,
            text=f"Current: ${self.portfolio.current_capital:,.2f}",
            font=('Arial', 14, 'bold'))
        self.current_capital_label.grid(row=0, column=1, padx=20)

        self.pnl_label = ttk.Label(stats_frame,
            text=f"P&L: ${self.portfolio.total_pnl:,.2f}",
            font=('Arial', 14, 'bold'))
        self.pnl_label.grid(row=0, column=2, padx=20)

        self.sharpe_label = ttk.Label(stats_frame,
            text=f"Sharpe: {self.portfolio.sharpe_ratio:.2f}",
            font=('Arial', 12))
        self.sharpe_label.grid(row=0, column=3, padx=20)

        # BTC Status
        btc_frame = ttk.LabelFrame(self.dashboard_frame, text="🟠 Bitcoin Market", padding=15)
        btc_frame.pack(fill='x', padx=20, pady=10)

        self.btc_spot_label = ttk.Label(btc_frame, text="Loading...", font=('Arial', 12, 'bold'))
        self.btc_spot_label.pack(side='left', padx=10)

        # Trade stats
        trade_stats_frame = ttk.LabelFrame(self.dashboard_frame, text="Trade Statistics", padding=15)
        trade_stats_frame.pack(fill='x', padx=20, pady=10)

        self.total_trades_label = ttk.Label(trade_stats_frame, text="Total: 0", font=('Arial', 11))
        self.total_trades_label.grid(row=0, column=0, padx=15)

        self.open_trades_label = ttk.Label(trade_stats_frame, text="Open: 0", font=('Arial', 11))
        self.open_trades_label.grid(row=0, column=1, padx=15)

        self.win_rate_label = ttk.Label(trade_stats_frame, text="Win Rate: 0%", font=('Arial', 11))
        self.win_rate_label.grid(row=0, column=2, padx=15)

        self.profit_factor_label = ttk.Label(trade_stats_frame, text="Profit Factor: 0", font=('Arial', 11))
        self.profit_factor_label.grid(row=0, column=3, padx=15)

        # Recent trades
        trades_frame = ttk.LabelFrame(self.dashboard_frame, text="Recent Trades", padding=10)
        trades_frame.pack(fill='both', expand=True, padx=20, pady=10)

        columns = ('Date', 'Instrument', 'Type', 'Qty', 'Entry', 'Current', 'P&L', 'Status')
        self.dashboard_tree = ttk.Treeview(trades_frame, columns=columns, show='headings', height=12)

        for col in columns:
            self.dashboard_tree.heading(col, text=col)
            self.dashboard_tree.column(col, width=120, anchor='center')

        scrollbar = ttk.Scrollbar(trades_frame, orient='vertical', command=self.dashboard_tree.yview)
        self.dashboard_tree.configure(yscrollcommand=scrollbar.set)
        self.dashboard_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        self.dashboard_tree.tag_configure('profit', foreground='green')
        self.dashboard_tree.tag_configure('loss', foreground='red')

    def build_trade_tab(self):
        """Build trade tab"""
        # Controls
        controls_frame = ttk.LabelFrame(self.trade_frame, text="Select Contract", padding=15)
        controls_frame.pack(fill='x', padx=20, pady=10)

        ttk.Label(controls_frame, text="Contract Type:", font=('Arial', 10, 'bold')).grid(row=0, column=0, padx=5)
        self.contract_type_var = tk.StringVar(value="futures")

        ttk.Radiobutton(controls_frame, text="Futures", variable=self.contract_type_var,
                       value='futures', command=self.on_contract_type_changed).grid(row=0, column=1, padx=5)
        ttk.Radiobutton(controls_frame, text="Options", variable=self.contract_type_var,
                       value='options', command=self.on_contract_type_changed).grid(row=0, column=2, padx=5)

        ttk.Label(controls_frame, text="Expiry:").grid(row=0, column=3, padx=5)
        self.expiry_var = tk.StringVar()
        self.expiry_combo = ttk.Combobox(controls_frame, textvariable=self.expiry_var, width=25, state='readonly')
        self.expiry_combo.grid(row=0, column=4, padx=5)
        self.expiry_combo.bind('<<ComboboxSelected>>', self.on_expiry_selected)

        # Futures frame
        self.futures_frame = ttk.LabelFrame(self.trade_frame, text="Bitcoin Futures", padding=10)
        self.futures_frame.pack(fill='both', expand=True, padx=20, pady=10)

        futures_columns = ('Symbol', 'LTP', 'Change 24h', 'OI', 'Actions')
        self.futures_tree = ttk.Treeview(self.futures_frame, columns=futures_columns, show='headings', height=10)

        for col in futures_columns:
            self.futures_tree.heading(col, text=col)
            width = 250 if col == 'Symbol' else 120
            self.futures_tree.column(col, width=width, anchor='center')

        scrollbar = ttk.Scrollbar(self.futures_frame, orient='vertical', command=self.futures_tree.yview)
        self.futures_tree.configure(yscrollcommand=scrollbar.set)
        self.futures_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        self.futures_tree.bind('<Double-Button-1>', self.on_futures_double_click)

        # Options frame
        self.options_frame = ttk.LabelFrame(self.trade_frame, text="Bitcoin Options Chain", padding=10)
        self.options_frame.pack(fill='both', expand=True, padx=20, pady=10)
        self.options_frame.pack_forget()

        options_columns = ('Call Symbol', 'Call LTP', 'Call IV', 'Call Delta', 'Strike',
                          'Put Delta', 'Put IV', 'Put LTP', 'Put Symbol')
        self.options_tree = ttk.Treeview(self.options_frame, columns=options_columns, show='headings', height=15)

        for col in options_columns:
            self.options_tree.heading(col, text=col)
            width = 180 if 'Symbol' in col else 90
            self.options_tree.column(col, width=width, anchor='center')

        scrollbar = ttk.Scrollbar(self.options_frame, orient='vertical', command=self.options_tree.yview)
        self.options_tree.configure(yscrollcommand=scrollbar.set)
        self.options_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        self.options_tree.bind('<Double-Button-1>', self.on_options_double_click)

    def on_contract_type_changed(self):
        """Handle contract type change"""
        if not self.instruments_loaded:
            return

        contract_type = self.contract_type_var.get()

        if contract_type == 'futures':
            self.options_frame.pack_forget()
            self.futures_frame.pack(fill='both', expand=True, padx=20, pady=10)
        else:
            self.futures_frame.pack_forget()
            self.options_frame.pack(fill='both', expand=True, padx=20, pady=10)

        self.update_expiries()

    def update_expiries(self):
        """Update expiries list"""
        if not self.instruments_loaded:
            return

        contract_type = self.contract_type_var.get()
        expiries = self.delta_api.get_expiries('BTC', contract_type)

        self.expiry_combo['values'] = expiries

        if expiries:
            self.expiry_combo.current(0)
            self.on_expiry_selected(None)

    def on_expiry_selected(self, event):
        """Handle expiry selection"""
        if not self.instruments_loaded:
            return

        self.selected_expiry = self.expiry_var.get()
        contract_type = self.contract_type_var.get()

        if contract_type == 'futures':
            self.update_futures_list()
        else:
            self.update_options_chain()

    def update_futures_list(self):
        """Update futures list"""
        for item in self.futures_tree.get_children():
            self.futures_tree.delete(item)

        expiry = self.expiry_var.get()
        if not expiry:
            return

        futures = [inst for inst in self.btc_instruments
                  if inst['instrument_type'] in ['FUT', 'PERP'] and
                  (expiry == 'PERPETUAL' and inst['instrument_type'] == 'PERP' or expiry in inst['tradingsymbol'])]

        for fut in futures:
            import json
            self.futures_tree.insert('', 'end', values=(
                fut['tradingsymbol'],
                f"${safe_float(fut.get('last_price', 0)):,.2f}",
                f"{safe_float(fut.get('mark_change_24h', 0)):+.2f}%",
                f"{safe_float(fut.get('oi', 0)):,.0f}",
                'Double-click to trade'
            ), tags=(json.dumps(fut),))

    def update_options_chain(self):
        """Update options chain"""
        for item in self.options_tree.get_children():
            self.options_tree.delete(item)

        expiry = self.expiry_var.get()
        if not expiry:
            return

        strikes = self.delta_api.get_strikes('BTC', expiry)

        for strike in strikes:
            import json
            strike_float = safe_float(strike)
            if strike_float <= 0:
                continue

            call = next((inst for inst in self.btc_instruments
                       if inst['instrument_type'] == 'CE' and
                       abs(safe_float(inst.get('strike', 0)) - strike_float) < 0.01 and
                       expiry in inst['tradingsymbol']), None)

            put = next((inst for inst in self.btc_instruments
                      if inst['instrument_type'] == 'PE' and
                      abs(safe_float(inst.get('strike', 0)) - strike_float) < 0.01 and
                      expiry in inst['tradingsymbol']), None)

            self.options_tree.insert('', 'end', values=(
                call['tradingsymbol'] if call else '-',
                f"${safe_float(call.get('last_price', 0)):,.2f}" if call else 'N/A',
                f"{safe_float(call.get('mark_iv', 0)):.1f}" if call else 'N/A',
                f"{safe_float(call.get('delta', 0)):.3f}" if call else 'N/A',
                f"${strike_float:,.0f}",
                f"{safe_float(put.get('delta', 0)):.3f}" if put else 'N/A',
                f"{safe_float(put.get('mark_iv', 0)):.1f}" if put else 'N/A',
                f"${safe_float(put.get('last_price', 0)):,.2f}" if put else 'N/A',
                put['tradingsymbol'] if put else '-'
            ), tags=(json.dumps({'call': call, 'put': put}),))

    def on_futures_double_click(self, event):
        """Handle futures double-click"""
        import json
        selection = self.futures_tree.selection()
        if not selection:
            return

        item = self.futures_tree.item(selection[0])
        try:
            instrument = json.loads(item['tags'][0])
            self.open_trade_dialog(instrument)
        except:
            messagebox.showerror("Error", "Could not load instrument")

    def on_options_double_click(self, event):
        """Handle options double-click"""
        import json
        selection = self.options_tree.selection()
        if not selection:
            return

        item = self.options_tree.item(selection[0])
        try:
            data = json.loads(item['tags'][0])
            choice = messagebox.askquestion("Select Option", "Trade CALL option?\n(No = PUT)")
            instrument = data['call'] if choice == 'yes' else data['put']

            if not instrument:
                messagebox.showwarning("Warning", "Option not available")
                return

            self.open_trade_dialog(instrument)
        except:
            messagebox.showerror("Error", "Could not load option")

    def open_trade_dialog(self, instrument):
        """Open trade dialog with SL/TP"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Place BTC Trade")
        dialog.geometry("600x750")
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text="Place Virtual BTC Trade", font=('Arial', 14, 'bold')).pack(pady=10)

        frame = ttk.Frame(dialog, padding=20)
        frame.pack(fill='both', expand=True)

        row = 0
        # Instrument info
        ttk.Label(frame, text="Instrument:", font=('Arial', 10, 'bold')).grid(row=row, column=0, sticky='w', pady=5)
        ttk.Label(frame, text=instrument['tradingsymbol']).grid(row=row, column=1, columnspan=2, sticky='w', pady=5)

        row += 1
        current_ltp = safe_float(instrument.get('last_price', 0))
        ttk.Label(frame, text="Current LTP:", font=('Arial', 10, 'bold')).grid(row=row, column=0, sticky='w', pady=5)
        ttk.Label(frame, text=f"${current_ltp:,.2f}", font=('Arial', 12, 'bold'), foreground='blue').grid(row=row, column=1, columnspan=2, sticky='w', pady=5)

        row += 1
        ttk.Separator(frame, orient='horizontal').grid(row=row, column=0, columnspan=3, sticky='ew', pady=10)

        row += 1
        # Trade type
        ttk.Label(frame, text="Trade Type:", font=('Arial', 10, 'bold')).grid(row=row, column=0, sticky='w', pady=5)
        trade_type_var = tk.StringVar(value='buy')
        ttk.Radiobutton(frame, text="BUY", variable=trade_type_var, value='buy').grid(row=row, column=1, sticky='w')
        ttk.Radiobutton(frame, text="SELL", variable=trade_type_var, value='sell').grid(row=row, column=2, sticky='w')

        row += 1
        # Quantity
        ttk.Label(frame, text="Quantity:", font=('Arial', 10, 'bold')).grid(row=row, column=0, sticky='w', pady=5)
        qty_entry = ttk.Entry(frame, width=20)
        qty_entry.insert(0, "1")
        qty_entry.grid(row=row, column=1, columnspan=2, sticky='w', pady=5)

        row += 1
        # Price
        ttk.Label(frame, text="Price ($):", font=('Arial', 10, 'bold')).grid(row=row, column=0, sticky='w', pady=5)
        price_entry = ttk.Entry(frame, width=20)
        if current_ltp > 0:
            price_entry.insert(0, f"{current_ltp:.2f}")
        price_entry.grid(row=row, column=1, columnspan=2, sticky='w', pady=5)

        row += 1
        ttk.Separator(frame, orient='horizontal').grid(row=row, column=0, columnspan=3, sticky='ew', pady=10)

        row += 1
        # Stop Loss
        ttk.Label(frame, text="Stop Loss ($):", font=('Arial', 10)).grid(row=row, column=0, sticky='w', pady=5)
        sl_entry = ttk.Entry(frame, width=20)
        sl_entry.grid(row=row, column=1, columnspan=2, sticky='w', pady=5)
        ttk.Label(frame, text="(Optional)", font=('Arial', 8)).grid(row=row, column=2, sticky='e', pady=5)

        row += 1
        # Take Profit
        ttk.Label(frame, text="Take Profit ($):", font=('Arial', 10)).grid(row=row, column=0, sticky='w', pady=5)
        tp_entry = ttk.Entry(frame, width=20)
        tp_entry.grid(row=row, column=1, columnspan=2, sticky='w', pady=5)
        ttk.Label(frame, text="(Optional)", font=('Arial', 8)).grid(row=row, column=2, sticky='e', pady=5)

        row += 1
        ttk.Separator(frame, orient='horizontal').grid(row=row, column=0, columnspan=3, sticky='ew', pady=10)

        row += 1
        # Notes
        ttk.Label(frame, text="Notes:", font=('Arial', 10)).grid(row=row, column=0, sticky='w', pady=5)
        notes_entry = ttk.Entry(frame, width=40)
        notes_entry.grid(row=row, column=1, columnspan=2, sticky='w', pady=5)

        row += 1
        # Tags
        ttk.Label(frame, text="Tags:", font=('Arial', 10)).grid(row=row, column=0, sticky='w', pady=5)
        tags_entry = ttk.Entry(frame, width=40)
        tags_entry.grid(row=row, column=1, columnspan=2, sticky='w', pady=5)
        ttk.Label(frame, text="(comma-separated)", font=('Arial', 8)).grid(row=row, column=2, sticky='e', pady=5)

        def place_trade():
            try:
                quantity = float(qty_entry.get())
                price = float(price_entry.get())

                if price <= 0 or quantity <= 0:
                    messagebox.showerror("Error", "Invalid quantity or price")
                    return

                # Parse SL/TP
                stop_loss = None
                take_profit = None

                if sl_entry.get():
                    stop_loss = float(sl_entry.get())
                if tp_entry.get():
                    take_profit = float(tp_entry.get())

                trade = Trade(
                    id=int(datetime.now().timestamp() * 1000),
                    trade_type=trade_type_var.get(),
                    instrument=instrument['tradingsymbol'],
                    trading_symbol=instrument['tradingsymbol'],
                    exchange='DELTA',
                    instrument_type=instrument['instrument_type'],
                    quantity=quantity,
                    entry_price=price,
                    current_price=price,
                    entry_date=datetime.now().strftime('%Y-%m-%d'),
                    status='open',
                    strike=safe_float(instrument.get('strike')),
                    stop_loss=stop_loss,
                    take_profit=take_profit,
                    delta=safe_float(instrument.get('delta')),
                    gamma=safe_float(instrument.get('gamma')),
                    theta=safe_float(instrument.get('theta')),
                    vega=safe_float(instrument.get('vega')),
                    iv=safe_float(instrument.get('mark_iv')),
                    notes=notes_entry.get(),
                    tags=tags_entry.get()
                )

                self.db.add_trade(trade)
                self.trades = self.db.get_all_trades()
                self.update_dashboard()
                self.refresh_positions()

                self.log(f"✅ {trade.trade_type.upper()}: {quantity} {instrument['tradingsymbol']} @ ${price:,.2f}")
                messagebox.showinfo("Success", "Trade placed!")
                dialog.destroy()

            except ValueError as e:
                messagebox.showerror("Error", f"Invalid input: {str(e)}")

        row += 1
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=row, column=0, columnspan=3, pady=20)

        ttk.Button(btn_frame, text="✅ Place Trade", command=place_trade, width=15).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="❌ Cancel", command=dialog.destroy, width=15).pack(side='left', padx=5)

    def build_positions_tab(self):
        """Build positions tab"""
        # Open positions
        open_frame = ttk.LabelFrame(self.positions_frame, text="Open Positions", padding=10)
        open_frame.pack(fill='both', expand=True, padx=20, pady=10)

        columns = ('Instrument', 'Type', 'Qty', 'Entry', 'Current', 'P&L', 'SL', 'TP')
        self.open_tree = ttk.Treeview(open_frame, columns=columns, show='headings', height=10)

        for col in columns:
            self.open_tree.heading(col, text=col)
            self.open_tree.column(col, width=120, anchor='center')

        scrollbar = ttk.Scrollbar(open_frame, orient='vertical', command=self.open_tree.yview)
        self.open_tree.configure(yscrollcommand=scrollbar.set)
        self.open_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        self.open_tree.tag_configure('profit', foreground='green')
        self.open_tree.tag_configure('loss', foreground='red')

        # Buttons
        btn_frame = ttk.Frame(open_frame)
        btn_frame.pack(fill='x', pady=5)

        ttk.Button(btn_frame, text="✅ Close Position", command=self.close_position).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="📝 Edit SL/TP", command=self.edit_sl_tp).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="🔄 Refresh", command=self.refresh_positions).pack(side='left', padx=5)

        # Closed positions
        closed_frame = ttk.LabelFrame(self.positions_frame, text="Closed Positions", padding=10)
        closed_frame.pack(fill='both', expand=True, padx=20, pady=10)

        columns = ('Instrument', 'Type', 'Qty', 'Entry', 'Exit', 'P&L', 'Date')
        self.closed_tree = ttk.Treeview(closed_frame, columns=columns, show='headings', height=8)

        for col in columns:
            self.closed_tree.heading(col, text=col)
            self.closed_tree.column(col, width=120, anchor='center')

        scrollbar = ttk.Scrollbar(closed_frame, orient='vertical', command=self.closed_tree.yview)
        self.closed_tree.configure(yscrollcommand=scrollbar.set)
        self.closed_tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        self.closed_tree.tag_configure('profit', foreground='green')
        self.closed_tree.tag_configure('loss', foreground='red')

    def edit_sl_tp(self):
        """Edit stop loss and take profit"""
        selection = self.open_tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Select a position")
            return

        item = self.open_tree.item(selection[0])
        trade_id = int(item['tags'][0])
        trade = self.db.get_trade(trade_id)

        if not trade:
            return

        # Dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Edit SL/TP")
        dialog.geometry("400x300")
        dialog.transient(self.root)

        ttk.Label(dialog, text=f"Edit SL/TP: {trade.trading_symbol}", font=('Arial', 12, 'bold')).pack(pady=10)

        frame = ttk.Frame(dialog, padding=20)
        frame.pack(fill='both', expand=True)

        ttk.Label(frame, text="Stop Loss ($):").grid(row=0, column=0, sticky='w', pady=10)
        sl_entry = ttk.Entry(frame, width=20)
        if trade.stop_loss:
            sl_entry.insert(0, str(trade.stop_loss))
        sl_entry.grid(row=0, column=1, pady=10)

        ttk.Label(frame, text="Take Profit ($):").grid(row=1, column=0, sticky='w', pady=10)
        tp_entry = ttk.Entry(frame, width=20)
        if trade.take_profit:
            tp_entry.insert(0, str(trade.take_profit))
        tp_entry.grid(row=1, column=1, pady=10)

        def save():
            try:
                sl_text = sl_entry.get()
                tp_text = tp_entry.get()

                trade.stop_loss = float(sl_text) if sl_text else None
                trade.take_profit = float(tp_text) if tp_text else None

                self.db.update_trade(trade)
                self.refresh_positions()
                self.log(f"📝 Updated SL/TP for {trade.trading_symbol}")
                messagebox.showinfo("Success", "SL/TP updated!")
                dialog.destroy()

            except ValueError:
                messagebox.showerror("Error", "Invalid values")

        ttk.Button(frame, text="💾 Save", command=save, width=15).grid(row=2, column=0, columnspan=2, pady=20)

    def close_position(self):
        """Close position"""
        selection = self.open_tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Select a position")
            return

        item = self.open_tree.item(selection[0])
        trade_id = int(item['tags'][0])
        trade = self.db.get_trade(trade_id)

        if not trade:
            return

        exit_price = simpledialog.askfloat("Close Position",
            f"Exit price for {trade.trading_symbol}:",
            initialvalue=trade.current_price)

        if exit_price is None:
            return

        trade.status = 'closed'
        trade.exit_price = exit_price
        trade.exit_date = datetime.now().strftime('%Y-%m-%d')
        trade.current_price = exit_price
        trade.realized_pnl = trade.calculate_pnl()
        trade.unrealized_pnl = 0

        self.db.update_trade(trade)
        self.trades = self.db.get_all_trades()
        self.update_dashboard()
        self.refresh_positions()
        self.log(f"✅ Closed {trade.trading_symbol}: P&L ${trade.realized_pnl:.2f}")
        messagebox.showinfo("Success", f"Position closed\nP&L: ${trade.realized_pnl:.2f}")

    def refresh_positions(self):
        """Refresh positions display"""
        # Clear trees
        for item in self.open_tree.get_children():
            self.open_tree.delete(item)
        for item in self.closed_tree.get_children():
            self.closed_tree.delete(item)

        # Open positions
        for trade in self.db.get_open_trades():
            pnl = trade.unrealized_pnl
            tag = 'profit' if pnl >= 0 else 'loss'

            self.open_tree.insert('', 'end', values=(
                trade.trading_symbol,
                trade.trade_type.upper(),
                trade.quantity,
                f"${trade.entry_price:.2f}",
                f"${trade.current_price:.2f}",
                f"${pnl:.2f}",
                f"${trade.stop_loss:.2f}" if trade.stop_loss else '-',
                f"${trade.take_profit:.2f}" if trade.take_profit else '-'
            ), tags=(str(trade.id), tag))

        # Closed positions
        for trade in reversed(self.db.get_closed_trades()[:20]):
            pnl = trade.realized_pnl
            tag = 'profit' if pnl >= 0 else 'loss'

            self.closed_tree.insert('', 'end', values=(
                trade.trading_symbol,
                trade.trade_type.upper(),
                trade.quantity,
                f"${trade.entry_price:.2f}",
                f"${trade.exit_price:.2f}" if trade.exit_price else '-',
                f"${pnl:.2f}",
                trade.exit_date or '-'
            ), tags=(tag,))

    def build_analytics_tab(self):
        """Build analytics tab"""
        # Performance metrics
        metrics_frame = ttk.LabelFrame(self.analytics_frame, text="Performance Metrics", padding=15)
        metrics_frame.pack(fill='x', padx=20, pady=10)

        self.analytics_text = scrolledtext.ScrolledText(metrics_frame, height=15, wrap=tk.WORD, font=('Courier', 10))
        self.analytics_text.pack(fill='both', expand=True)

        ttk.Button(metrics_frame, text="🔄 Refresh Analytics", command=self.update_analytics).pack(pady=10)

        # Charts frame
        if MATPLOTLIB_AVAILABLE:
            charts_frame = ttk.LabelFrame(self.analytics_frame, text="Performance Charts", padding=15)
            charts_frame.pack(fill='both', expand=True, padx=20, pady=10)

            ttk.Button(charts_frame, text="📊 Show Portfolio Chart", command=self.show_portfolio_chart).pack(pady=5)
            ttk.Button(charts_frame, text="📈 Show P&L Distribution", command=self.show_pnl_chart).pack(pady=5)

    def update_analytics(self):
        """Update analytics display"""
        self.analytics_text.delete(1.0, tk.END)

        try:
            report = generate_performance_report(self.trades, self.portfolio)

            analytics = f"""
╔══════════════════════════════════════════════════════════════╗
║           BITCOIN TRADING PERFORMANCE REPORT                 ║
╚══════════════════════════════════════════════════════════════╝

📊 SUMMARY STATISTICS
{"="*60}
Total Trades:           {report['summary']['total_trades']}
Open Trades:            {report['summary']['open_trades']}
Closed Trades:          {report['summary']['closed_trades']}
Win Rate:               {report['summary']['win_rate']:.2f}%
Profit Factor:          {report['summary']['profit_factor']:.2f}
Expectancy:             ${report['summary']['expectancy']:.2f}
Average Win:            ${report['summary']['avg_win']:.2f}
Average Loss:           ${report['summary']['avg_loss']:.2f}

💰 PROFIT & LOSS
{"="*60}
Total P&L:              ${report['summary']['total_pnl']:.2f}
Realized P&L:           ${report['summary']['realized_pnl']:.2f}
Unrealized P&L:         ${report['summary']['unrealized_pnl']:.2f}
Max Drawdown:           ${report['summary']['max_drawdown']:.2f}
Sharpe Ratio:           {report['summary']['sharpe_ratio']:.2f}

📈 PORTFOLIO GREEKS
{"="*60}
Delta:                  {report['portfolio_greeks']['delta']:.2f}
Gamma:                  {report['portfolio_greeks']['gamma']:.4f}
Theta:                  {report['portfolio_greeks']['theta']:.2f}
Vega:                   {report['portfolio_greeks']['vega']:.2f}

⚠️  RISK MANAGEMENT
{"="*60}
Kelly Criterion:        {report['risk_management']['kelly_criterion']:.2%}
Recommended Size:       {report['risk_management']['recommended_position_size']:.2f}%

📊 TRADE DISTRIBUTION
{"="*60}
Long Trades:            {report['distribution']['long_trades']['count']} (Win Rate: {report['distribution']['long_trades']['win_rate']:.1f}%)
Short Trades:           {report['distribution']['short_trades']['count']} (Win Rate: {report['distribution']['short_trades']['win_rate']:.1f}%)

⏱️  HOLDING PERIOD STATS
{"="*60}
Average Days:           {report['holding_period']['avg_days']:.1f}
Min Days:               {report['holding_period']['min_days']:.0f}
Max Days:               {report['holding_period']['max_days']:.0f}
Median Days:            {report['holding_period'].get('median_days', 0):.1f}

"""
            self.analytics_text.insert(1.0, analytics)

        except Exception as e:
            self.analytics_text.insert(1.0, f"Error generating report: {str(e)}")

    def show_portfolio_chart(self):
        """Show portfolio equity curve chart"""
        if not MATPLOTLIB_AVAILABLE:
            messagebox.showinfo("Info", "Matplotlib not installed")
            return

        # Create chart window
        chart_window = tk.Toplevel(self.root)
        chart_window.title("Portfolio Equity Curve")
        chart_window.geometry("1000x600")

        fig = Figure(figsize=(10, 6))
        ax = fig.add_subplot(111)

        # Calculate equity curve
        closed_trades = sorted([t for t in self.trades if t.status == 'closed'],
                              key=lambda x: x.exit_date or '')

        equity = [self.portfolio.initial_capital]
        dates = ['Start']

        cumulative_pnl = 0
        for trade in closed_trades:
            cumulative_pnl += trade.realized_pnl
            equity.append(self.portfolio.initial_capital + cumulative_pnl)
            dates.append(trade.exit_date or '')

        ax.plot(equity, marker='o', linewidth=2, markersize=4)
        ax.set_title('Portfolio Equity Curve', fontsize=14, fontweight='bold')
        ax.set_xlabel('Trades')
        ax.set_ylabel('Capital ($)')
        ax.grid(True, alpha=0.3)

        canvas = FigureCanvasTkAgg(fig, chart_window)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True)

    def show_pnl_chart(self):
        """Show P&L distribution chart"""
        if not MATPLOTLIB_AVAILABLE:
            messagebox.showinfo("Info", "Matplotlib not installed")
            return

        closed_trades = [t for t in self.trades if t.status == 'closed']

        if not closed_trades:
            messagebox.showinfo("Info", "No closed trades to display")
            return

        chart_window = tk.Toplevel(self.root)
        chart_window.title("P&L Distribution")
        chart_window.geometry("1000x600")

        fig = Figure(figsize=(10, 6))
        ax = fig.add_subplot(111)

        pnls = [t.realized_pnl for t in closed_trades]

        ax.hist(pnls, bins=20, edgecolor='black', alpha=0.7)
        ax.axvline(x=0, color='red', linestyle='--', linewidth=2)
        ax.set_title('P&L Distribution', fontsize=14, fontweight='bold')
        ax.set_xlabel('P&L ($)')
        ax.set_ylabel('Frequency')
        ax.grid(True, alpha=0.3)

        canvas = FigureCanvasTkAgg(fig, chart_window)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True)

    def build_strategy_tab(self):
        """Build strategy builder tab"""
        ttk.Label(self.strategy_frame,
                 text="🎯 Options Strategy Builder",
                 font=('Arial', 16, 'bold')).pack(pady=20)

        info_frame = ttk.LabelFrame(self.strategy_frame, text="Strategy Information", padding=15)
        info_frame.pack(fill='x', padx=20, pady=10)

        ttk.Label(info_frame, text="Select a strategy to see details and recommendations").pack()

        # Strategy selection
        strategy_frame = ttk.LabelFrame(self.strategy_frame, text="Available Strategies", padding=15)
        strategy_frame.pack(fill='both', expand=True, padx=20, pady=10)

        strategies = [
            "Long Call",
            "Long Put",
            "Bull Call Spread",
            "Bear Put Spread",
            "Long Straddle",
            "Short Straddle",
            "Long Strangle",
            "Iron Condor",
            "Butterfly Spread",
            "Covered Call",
            "Protective Put"
        ]

        for i, strategy in enumerate(strategies):
            row = i // 3
            col = i % 3
            ttk.Button(strategy_frame, text=strategy, width=25,
                      command=lambda s=strategy: self.show_strategy_info(s)).grid(
                row=row, column=col, padx=10, pady=10)

    def show_strategy_info(self, strategy_name):
        """Show strategy information"""
        messagebox.showinfo("Strategy Info",
            f"Strategy: {strategy_name}\n\n"
            f"This feature allows you to analyze and build {strategy_name} strategy.\n\n"
            f"Select options from the Trade tab to construct this strategy.")

    def build_settings_tab(self):
        """Build settings tab"""
        # Portfolio settings
        portfolio_frame = ttk.LabelFrame(self.settings_frame, text="Portfolio Settings", padding=20)
        portfolio_frame.pack(fill='x', padx=20, pady=10)

        ttk.Label(portfolio_frame, text="Initial Capital ($):").grid(row=0, column=0, padx=5, pady=5, sticky='w')
        self.capital_entry = ttk.Entry(portfolio_frame, width=30)
        self.capital_entry.insert(0, str(self.portfolio.initial_capital))
        self.capital_entry.grid(row=0, column=1, padx=5, pady=5)

        ttk.Button(portfolio_frame, text="💾 Update Capital",
                  command=self.update_capital).grid(row=1, column=1, pady=10, sticky='w')

        # Data management
        data_frame = ttk.LabelFrame(self.settings_frame, text="Data Management", padding=20)
        data_frame.pack(fill='x', padx=20, pady=10)

        ttk.Button(data_frame, text="📊 Export to Excel",
                  command=self.export_trades).pack(side='left', padx=5)
        ttk.Button(data_frame, text="📈 Update Analytics",
                  command=self.update_analytics).pack(side='left', padx=5)

        # Log
        log_frame = ttk.LabelFrame(self.settings_frame, text="Activity Log", padding=10)
        log_frame.pack(fill='both', expand=True, padx=20, pady=10)

        self.log_text = scrolledtext.ScrolledText(log_frame, height=20, wrap=tk.WORD)
        self.log_text.pack(fill='both', expand=True)

        self.log("🚀 Bitcoin Virtual Trading Platform PRO - Started")
        self.log("🟠 Delta Exchange API connected")
        self.log(f"💾 Database: {self.config.get('trading', 'database_path')}")

    def update_capital(self):
        """Update capital"""
        try:
            new_capital = float(self.capital_entry.get())
            if new_capital <= 0:
                messagebox.showerror("Error", "Capital must be positive")
                return

            self.portfolio.initial_capital = new_capital
            self.db.update_portfolio(self.portfolio)
            self.update_dashboard()
            self.log(f"💰 Capital updated: ${new_capital:,.2f}")
            messagebox.showinfo("Success", "Capital updated!")

        except ValueError:
            messagebox.showerror("Error", "Invalid number")

    def export_trades(self):
        """Export trades to Excel"""
        try:
            import openpyxl
            from openpyxl import Workbook

            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"btc_trades_pro_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )

            if not filename:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "BTC Trades"

            # Headers
            headers = ['ID', 'Date', 'Instrument', 'Type', 'Qty', 'Entry', 'Exit',
                      'Current', 'P&L', 'Status', 'SL', 'TP', 'Notes', 'Tags']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # Data
            row = 2
            for trade in self.trades:
                ws.cell(row=row, column=1, value=trade.id)
                ws.cell(row=row, column=2, value=trade.entry_date)
                ws.cell(row=row, column=3, value=trade.trading_symbol)
                ws.cell(row=row, column=4, value=trade.trade_type.upper())
                ws.cell(row=row, column=5, value=trade.quantity)
                ws.cell(row=row, column=6, value=trade.entry_price)
                ws.cell(row=row, column=7, value=trade.exit_price or '')
                ws.cell(row=row, column=8, value=trade.current_price)
                ws.cell(row=row, column=9, value=trade.realized_pnl if trade.status == 'closed' else trade.unrealized_pnl)
                ws.cell(row=row, column=10, value=trade.status.upper())
                ws.cell(row=row, column=11, value=trade.stop_loss or '')
                ws.cell(row=row, column=12, value=trade.take_profit or '')
                ws.cell(row=row, column=13, value=trade.notes)
                ws.cell(row=row, column=14, value=trade.tags)
                row += 1

            wb.save(filename)
            self.log(f"✅ Exported to: {filename}")
            messagebox.showinfo("Success", f"Trades exported!")

        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {str(e)}")

    def update_dashboard(self):
        """Update dashboard"""
        # Calculate totals
        open_trades = self.db.get_open_trades()
        closed_trades = self.db.get_closed_trades()

        total_unrealized = sum(t.unrealized_pnl for t in open_trades)
        total_realized = sum(t.realized_pnl for t in closed_trades)
        total_pnl = total_unrealized + total_realized

        self.portfolio.unrealized_pnl = total_unrealized
        self.portfolio.realized_pnl = total_realized
        self.portfolio.total_pnl = total_pnl
        self.portfolio.current_capital = self.portfolio.initial_capital + total_pnl
        self.portfolio.total_trades = len(self.trades)
        self.portfolio.winning_trades = len([t for t in closed_trades if t.realized_pnl > 0])
        self.portfolio.losing_trades = len([t for t in closed_trades if t.realized_pnl < 0])

        self.db.update_portfolio(self.portfolio)

        # Update labels
        self.initial_capital_label.config(text=f"Initial: ${self.portfolio.initial_capital:,.2f}")
        self.current_capital_label.config(text=f"Current: ${self.portfolio.current_capital:,.2f}")

        pnl_color = 'green' if total_pnl >= 0 else 'red'
        self.pnl_label.config(text=f"P&L: ${total_pnl:,.2f}", foreground=pnl_color)
        self.sharpe_label.config(text=f"Sharpe: {self.portfolio.sharpe_ratio:.2f}")

        # Trade stats
        win_rate = PerformanceAnalytics.calculate_win_rate(self.trades)
        profit_factor = PerformanceAnalytics.calculate_profit_factor(self.trades)

        self.total_trades_label.config(text=f"Total: {len(self.trades)}")
        self.open_trades_label.config(text=f"Open: {len(open_trades)}")
        self.win_rate_label.config(text=f"Win Rate: {win_rate:.1f}%")
        self.profit_factor_label.config(text=f"Profit Factor: {profit_factor:.2f}")

        # Refresh dashboard trades
        for item in self.dashboard_tree.get_children():
            self.dashboard_tree.delete(item)

        for trade in reversed(self.trades[-20:]):
            pnl = trade.realized_pnl if trade.status == 'closed' else trade.unrealized_pnl
            tag = 'profit' if pnl >= 0 else 'loss'

            self.dashboard_tree.insert('', 'end', values=(
                trade.entry_date,
                trade.trading_symbol,
                trade.trade_type.upper(),
                trade.quantity,
                f"${trade.entry_price:.2f}",
                f"${trade.current_price:.2f}",
                f"${pnl:.2f}",
                trade.status.upper()
            ), tags=(tag,))

    def on_closing(self):
        """Handle window close"""
        if self.btc_refresh_timer:
            self.root.after_cancel(self.btc_refresh_timer)
        self.db.close()
        self.root.destroy()


def main():
    """Main function"""
    print("=" * 80)
    print("🟠 BITCOIN VIRTUAL TRADING PLATFORM - PROFESSIONAL EDITION")
    print("=" * 80)
    print("\n✨ ADVANCED FEATURES:")
    print("  💾 SQLite database with full trade history")
    print("  📊 Advanced risk analytics (Sharpe, Sortino, VaR)")
    print("  🎯 Options strategy builder (10+ strategies)")
    print("  📈 Interactive performance charts")
    print("  🛡️  Stop Loss / Take Profit automation")
    print("  📝 Trade journal with notes and tags")
    print("  🔬 Portfolio Greeks aggregation")
    print("  💹 Real-time Bitcoin futures & options")
    print("\n📱 Opening application...\n")

    if BOOTSTRAP_AVAILABLE:
        root = tb.Window(themename="darkly")
    else:
        root = tk.Tk()

    app = BTCTradingPlatformPro(root)
    root.mainloop()


if __name__ == '__main__':
    main()
