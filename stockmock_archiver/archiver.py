#!/usr/bin/env python3
"""
StockMock Nifty option-chain archiver.

Pulls the StockMock nifty option chain for every trading weekday over a
chosen date range, writes it all to one Excel file, and (optionally)
opens it when done.

Upgrades over the original single-file script:

  * No secret in source. The auth token is never hardcoded -- it is read
    from --token, the STOCKMOCK_TOKEN environment variable, or an
    interactive prompt, in that order.
  * Real CLI (argparse) instead of editing constants at the top of the
    file. All the old module-level "CONFIG" knobs are now flags, with the
    same defaults as before.
  * Config is a single immutable-ish dataclass instead of scattered
    module globals, which also makes the core logic unit-testable.
  * Network retries use exponential backoff and understand HTTP 429
    (rate limiting), honoring a Retry-After header when the server sends
    one, instead of a single blind retry.
  * logging instead of print, so verbosity is controllable (-q/-v) and
    output is timestamped.

Behavioural notes carried over from the original script:

  * If the token expires mid-run (401/403), the run pauses and asks --
    right there in the terminal -- for a fresh token, then resumes the
    exact same snapshot. Nothing is lost.
  * Every (date, time) snapshot is cached to CSV as it's fetched, so a
    re-run after a stop (or a declined token refresh) picks up exactly
    where it left off.
  * NSE's Nifty weekly-expiry weekday has changed more than once over
    the years (historically Thursday, briefly Monday/Wednesday, Tuesday
    from late 2024 onward). "expiry_weekday" time mode is a heuristic
    based on weekday, not a guarantee -- adjust --expiry-weekdays to
    cover the regime(s) your date range spans.
  * Weekends are skipped automatically. Exchange holidays aren't known
    in advance -- they come back as "no data" for that date+time
    (recorded as such, not treated as an error).
"""

from __future__ import annotations

import argparse
import csv
import logging
import os
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Optional, Tuple

import requests
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

IST = timezone(timedelta(hours=5, minutes=30))
API_URL = "https://www.stockmock.in/api/getPlainAllOC"
CACHE_COLUMNS = [
    "request_date", "request_time", "index", "expiry", "spot", "future",
    "do", "gpt", "gpr", "strike", "CE_time", "CE_premium", "PE_time", "PE_premium",
]
VALID_TIME_MODES = ("single", "multi_all", "expiry_weekday")

log = logging.getLogger("stockmock_archiver")


# ============================== CONFIG ==============================

@dataclass
class Config:
    index: str = "nifty"

    # Date range. Leave both None to fall back to "last years_back years
    # from today".
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    years_back: int = 4

    # Time selection. See VALID_TIME_MODES / module docstring.
    time_mode: str = "expiry_weekday"
    selected_time: str = "09:16:00"
    intraday_start: str = "09:16:00"
    intraday_end: str = "15:29:00"
    intraday_interval_minutes: int = 15
    expiry_weekdays: List[int] = field(default_factory=lambda: [1, 3])  # Tue, Thu

    request_delay_seconds: float = 0.4
    max_retries: int = 3
    request_timeout_seconds: float = 20.0

    output_xlsx: str = "stockmock_option_chain.xlsx"
    cache_csv: str = "stockmock_cache.csv"
    auto_open: bool = True

    def validate(self) -> None:
        if self.time_mode not in VALID_TIME_MODES:
            raise ValueError(
                f"invalid time_mode {self.time_mode!r}, expected one of {VALID_TIME_MODES}"
            )
        if self.intraday_interval_minutes <= 0:
            raise ValueError("intraday_interval_minutes must be > 0")
        if not all(0 <= d <= 6 for d in self.expiry_weekdays):
            raise ValueError("expiry_weekdays entries must be 0 (Mon) .. 6 (Sun)")
        for label, value in (
            ("selected_time", self.selected_time),
            ("intraday_start", self.intraday_start),
            ("intraday_end", self.intraday_end),
        ):
            _parse_time(value, label)
        if _parse_time(self.intraday_start, "intraday_start") > _parse_time(
            self.intraday_end, "intraday_end"
        ):
            raise ValueError("intraday_start must not be after intraday_end")
        for label, value in (("start_date", self.start_date), ("end_date", self.end_date)):
            if value is not None:
                _parse_date(value, label)
        if self.start_date and self.end_date and self.start_date > self.end_date:
            raise ValueError(f"start_date ({self.start_date}) is after end_date ({self.end_date})")
        if self.years_back <= 0:
            raise ValueError("years_back must be > 0")
        if self.request_delay_seconds < 0:
            raise ValueError("request_delay_seconds must be >= 0")
        if self.max_retries < 1:
            raise ValueError("max_retries must be >= 1")


def _parse_time(value: str, label: str) -> datetime:
    try:
        return datetime.strptime(value, "%H:%M:%S")
    except ValueError as exc:
        raise ValueError(f"{label} must be HH:MM:SS, got {value!r}") from exc


def _parse_date(value: str, label: str) -> datetime:
    try:
        return datetime.strptime(value, "%Y-%m-%d")
    except ValueError as exc:
        raise ValueError(f"{label} must be YYYY-MM-DD, got {value!r}") from exc


class TokenBox:
    """Small mutable holder so a refreshed token is visible everywhere
    without reaching for a module-level global."""

    def __init__(self, token: str):
        self.token = token


def resolve_token(cli_token: Optional[str]) -> str:
    if cli_token:
        return cli_token
    env_token = os.environ.get("STOCKMOCK_TOKEN")
    if env_token:
        return env_token
    print(
        "No token supplied via --token or STOCKMOCK_TOKEN.\n"
        "Get one from: DevTools (F12) -> Network -> Fetch/XHR -> getPlainAllOC -> Headers -> 'token'."
    )
    token = input("Paste your StockMock token: ").strip()
    if not token:
        raise SystemExit("ERROR: no token provided -- cannot continue.")
    return token


def make_headers(token: str) -> Dict[str, str]:
    return {
        "Referer": "https://www.stockmock.in/",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/150.0.0.0 Safari/537.36"
        ),
        "Accept": "application/json, text/*",
        "Content-Type": "application/json",
        "feversion": "2",
        "DNT": "1",
        "token": token,
    }


def compute_st_date_ms(date_str: str, time_str: str) -> int:
    dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S").replace(tzinfo=IST)
    return int(dt.timestamp() * 1000)


def split_time_premium(val) -> Tuple[Optional[str], Optional[float]]:
    """'09:16-4.3' -> ('09:16', 4.3). 0 or None -> (None, None)."""
    if val in (0, "0", None):
        return None, None
    if isinstance(val, str) and "-" in val:
        t, p = val.rsplit("-", 1)
        try:
            p = float(p)
        except ValueError:
            pass
        return t, p
    return None, val


def parse_final_result(payload: dict, request_date: str, request_time: str, index: str) -> List[dict]:
    """
    Flatten one snapshot's response into a list of row dicts.
    Handles the real observed shape:
      {"finalResult": [ {expiry_label: {do,gpt,gpr,s,f,c:[[strike,ce,pe],...]}}, ... ] }
    A batch that comes back as an empty list (no data for that day/time)
    is treated as "nothing to record", not an error.
    """
    final_result = payload.get("finalResult", [])
    rows = []
    for batch in final_result:
        if not isinstance(batch, dict):
            continue  # empty list batch == no data
        for expiry_label, meta in batch.items():
            if not isinstance(meta, dict):
                continue
            do = meta.get("do")
            gpt = meta.get("gpt")
            gpr = meta.get("gpr")
            spot = meta.get("s")
            future = meta.get("f")
            for entry in meta.get("c", []):
                if not isinstance(entry, list) or len(entry) != 3:
                    continue
                strike, ce_raw, pe_raw = entry
                ce_time, ce_prem = split_time_premium(ce_raw)
                pe_time, pe_prem = split_time_premium(pe_raw)
                rows.append({
                    "request_date": request_date,
                    "request_time": request_time,
                    "index": index,
                    "expiry": expiry_label,
                    "spot": spot,
                    "future": future,
                    "do": do,
                    "gpt": gpt,
                    "gpr": gpr,
                    "strike": strike,
                    "CE_time": ce_time,
                    "CE_premium": ce_prem,
                    "PE_time": pe_time,
                    "PE_premium": pe_prem,
                })
    return rows


def generate_intraday_times(start: str, end: str, interval_minutes: int) -> List[str]:
    times = []
    t = datetime.strptime(start, "%H:%M:%S")
    end_t = datetime.strptime(end, "%H:%M:%S")
    while t <= end_t:
        times.append(t.strftime("%H:%M:%S"))
        t += timedelta(minutes=interval_minutes)
    return times


def times_for_date(date_str: str, config: Config, intraday_times: List[str]) -> List[str]:
    if config.time_mode == "single":
        return [config.selected_time]
    if config.time_mode == "multi_all":
        return intraday_times
    if config.time_mode == "expiry_weekday":
        weekday = datetime.strptime(date_str, "%Y-%m-%d").weekday()
        if weekday in config.expiry_weekdays:
            return intraday_times
        return [config.selected_time]
    raise ValueError(f"unknown time_mode {config.time_mode!r}")


def resolve_date_range(config: Config, today: Optional[datetime] = None) -> Tuple[datetime.date, datetime.date]:
    today = today or datetime.now(IST)
    if config.start_date or config.end_date:
        end_date = (
            datetime.strptime(config.end_date, "%Y-%m-%d").date()
            if config.end_date
            else today.date()
        )
        start_date = (
            datetime.strptime(config.start_date, "%Y-%m-%d").date()
            if config.start_date
            else end_date.replace(year=end_date.year - config.years_back)
        )
    else:
        end_date = today.date()
        start_date = end_date.replace(year=end_date.year - config.years_back)
    if start_date > end_date:
        raise ValueError(f"start_date ({start_date}) is after end_date ({end_date})")
    return start_date, end_date


def build_task_list(
    start_date, end_date, config: Config, intraday_times: List[str]
) -> List[Tuple[str, str]]:
    tasks = []
    d = start_date
    while d <= end_date:
        if d.weekday() < 5:  # Mon-Fri only; exchange holidays handled as "no data" responses
            date_str = d.strftime("%Y-%m-%d")
            for t in times_for_date(date_str, config, intraday_times):
                tasks.append((date_str, t))
        d += timedelta(days=1)
    return tasks


# ============================== NETWORK ==============================

def fetch_one_snapshot(
    session: requests.Session,
    token_box: TokenBox,
    config: Config,
    date_str: str,
    time_str: str,
) -> Tuple[str, List[dict]]:
    """Returns (status, rows). status is 'ok', 'empty', or 'auth_error'.

    Transient failures (network errors, 5xx, 429) are retried with
    exponential backoff up to config.max_retries attempts; 429 honors a
    Retry-After header when present.
    """
    body = {
        "index": config.index,
        "selectedDate": date_str,
        "stDate": compute_st_date_ms(date_str, time_str),
        "selectedTime": time_str,
        "int": 0.05,
        "simulatorBasePrice": "spot",
    }

    backoff = 2.0
    for attempt in range(1, config.max_retries + 1):
        try:
            resp = session.post(
                API_URL,
                headers=make_headers(token_box.token),
                json=body,
                timeout=config.request_timeout_seconds,
            )
        except requests.exceptions.RequestException as exc:
            log.warning("[%s %s] network error (attempt %d/%d): %s", date_str, time_str, attempt, config.max_retries, exc)
            if attempt == config.max_retries:
                return "empty", []
            time.sleep(backoff)
            backoff *= 2
            continue

        if resp.status_code in (401, 403):
            return "auth_error", []

        if resp.status_code == 429:
            retry_after = resp.headers.get("Retry-After")
            wait = float(retry_after) if retry_after and retry_after.isdigit() else backoff
            log.warning("[%s %s] rate limited (429), waiting %.1fs (attempt %d/%d)", date_str, time_str, wait, attempt, config.max_retries)
            if attempt == config.max_retries:
                return "empty", []
            time.sleep(wait)
            backoff *= 2
            continue

        if 500 <= resp.status_code < 600:
            log.warning("[%s %s] HTTP %d (attempt %d/%d): %s", date_str, time_str, resp.status_code, attempt, config.max_retries, resp.text[:200])
            if attempt == config.max_retries:
                return "empty", []
            time.sleep(backoff)
            backoff *= 2
            continue

        if resp.status_code != 200:
            log.warning("[%s %s] HTTP %d: %s -- skipping", date_str, time_str, resp.status_code, resp.text[:200])
            return "empty", []

        try:
            payload = resp.json()
        except ValueError:
            log.warning("[%s %s] non-JSON response -- skipping", date_str, time_str)
            return "empty", []

        rows = parse_final_result(payload, date_str, time_str, config.index)
        return ("ok" if rows else "empty"), rows

    return "empty", []


def fetch_snapshot_with_retry(
    session: requests.Session, token_box: TokenBox, config: Config, date_str: str, time_str: str
) -> Tuple[str, List[dict]]:
    """
    Wraps fetch_one_snapshot: on a 401/403, pauses and asks -- right here
    in the terminal -- for a fresh token, then retries the SAME snapshot.
    Returns (status, rows) where status is 'ok', 'empty', or 'abort'
    ('abort' means the user chose to stop instead of pasting a new token).
    """
    while True:
        status, rows = fetch_one_snapshot(session, token_box, config, date_str, time_str)
        if status != "auth_error":
            return status, rows

        print(f"\n[{date_str} {time_str}] HTTP 401/403 -- your token has expired.")
        print("Get a fresh one: DevTools (F12) -> Network -> Fetch/XHR -> getPlainAllOC -> Headers -> 'token'.")
        try:
            new_token = input(
                "Paste the new token here and press Enter (or press Enter alone to stop and save progress): "
            ).strip()
        except EOFError:
            new_token = ""

        if not new_token:
            return "abort", []

        token_box.token = new_token
        log.info("Token updated -- retrying this snapshot...")


# ============================== CACHE / EXCEL ==============================

def load_cached_keys(config: Config) -> set:
    """Returns the set of (request_date, request_time) pairs already cached."""
    if not os.path.exists(config.cache_csv):
        return set()
    try:
        df = pd.read_csv(config.cache_csv)
        if "request_time" in df.columns:
            req_times = df["request_time"].fillna(config.selected_time).astype(str)
        else:
            # Legacy cache written before multi-time support existed --
            # every row in it was fetched at selected_time.
            req_times = pd.Series([config.selected_time] * len(df))
        req_dates = df["request_date"].astype(str)
        return set(zip(req_dates, req_times))
    except Exception:
        return set()


def append_to_cache(config: Config, rows: List[dict]) -> None:
    file_exists = os.path.exists(config.cache_csv)
    with open(config.cache_csv, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CACHE_COLUMNS)
        if not file_exists:
            writer.writeheader()
        for row in rows:
            writer.writerow(row)


def mark_snapshot_done_with_no_data(config: Config, date_str: str, time_str: str) -> None:
    """Record an explicit 'no data' marker row so this snapshot is skipped on resume."""
    append_to_cache(config, [{
        "request_date": date_str, "request_time": time_str, "index": config.index, "expiry": "NO_DATA",
        "spot": None, "future": None, "do": None, "gpt": None, "gpr": None,
        "strike": None, "CE_time": None, "CE_premium": None,
        "PE_time": None, "PE_premium": None,
    }])


def build_excel_from_cache(config: Config) -> pd.DataFrame:
    if not os.path.exists(config.cache_csv):
        raise SystemExit("ERROR: no cache file found -- nothing was ever fetched.")

    df = pd.read_csv(config.cache_csv)
    real_rows = df[df["expiry"] != "NO_DATA"].copy()

    if real_rows.empty:
        raise SystemExit("ERROR: cache exists but contains no real option-chain rows -- nothing to write.")

    col_order = [c for c in CACHE_COLUMNS if c in real_rows.columns]
    real_rows = real_rows[col_order]

    with pd.ExcelWriter(config.output_xlsx, engine="openpyxl") as writer:
        real_rows.to_excel(writer, index=False, sheet_name="Option Chain")

        summary = pd.DataFrame({
            "Field": [
                "Index", "Total Rows", "Unique Dates With Data", "Unique Snapshot Times",
                "Time Mode", "Date Range Requested", "Generated At",
            ],
            "Value": [
                config.index,
                len(real_rows),
                real_rows["request_date"].nunique(),
                real_rows["request_time"].nunique() if "request_time" in real_rows.columns else 1,
                config.time_mode,
                f"{config.start_date or '(auto)'} to {config.end_date or '(auto)'}"
                if (config.start_date or config.end_date)
                else f"last {config.years_back} years",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ],
        })
        summary.to_excel(writer, index=False, sheet_name="Summary")

        ws = writer.sheets["Option Chain"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
        for col_idx, col_name in enumerate(real_rows.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            # NaN-safe stringification (astype(str) on real NaN floats can
            # leave a bare float in the list and crash len())
            sample = [str(v) if pd.notna(v) else "" for v in real_rows[col_name].tolist()[:200]]
            max_len = max([len(str(col_name))] + [len(s) for s in sample])
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)
        ws.freeze_panes = "A2"

    log.info("Excel written: %s (%d rows)", os.path.abspath(config.output_xlsx), len(real_rows))
    return real_rows


def auto_open(path: str) -> None:
    try:
        if sys.platform.startswith("win"):
            os.startfile(path)  # noqa
        elif sys.platform == "darwin":
            os.system(f'open "{path}"')
        else:
            os.system(f'xdg-open "{path}"')
    except Exception as exc:
        log.warning("Couldn't auto-open the file: %s. Open it manually from the path above.", exc)


# ============================== CLI / MAIN ==============================

def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Archive the StockMock Nifty option chain over a date range into one Excel file.",
    )
    parser.add_argument("--token", help="StockMock auth token. Falls back to STOCKMOCK_TOKEN env var, then an interactive prompt.")
    parser.add_argument("--index", default="nifty", help="Index to fetch (default: nifty).")
    parser.add_argument("--start-date", help="YYYY-MM-DD. Defaults to (end-date - years-back).")
    parser.add_argument("--end-date", help="YYYY-MM-DD. Defaults to today.")
    parser.add_argument("--years-back", type=int, default=4, help="Used when --start-date is omitted (default: 4).")
    parser.add_argument("--time-mode", choices=VALID_TIME_MODES, default="expiry_weekday")
    parser.add_argument("--selected-time", default="09:16:00", help="HH:MM:SS single snapshot time (default: 09:16:00).")
    parser.add_argument("--intraday-start", default="09:16:00")
    parser.add_argument("--intraday-end", default="15:29:00")
    parser.add_argument("--interval-minutes", type=int, default=15, help="Gap between intraday snapshots (default: 15).")
    parser.add_argument(
        "--expiry-weekdays", default="1,3",
        help="Comma-separated weekdays (Mon=0..Sun=6) treated as expiry days for --time-mode expiry_weekday (default: 1,3 = Tue,Thu).",
    )
    parser.add_argument("--delay", type=float, default=0.4, help="Seconds to sleep between requests (default: 0.4).")
    parser.add_argument("--max-retries", type=int, default=3, help="Retry attempts for transient failures (default: 3).")
    parser.add_argument("--output", default="stockmock_option_chain.xlsx", help="Output .xlsx path.")
    parser.add_argument("--cache", default="stockmock_cache.csv", help="Cache .csv path (resumable).")
    parser.add_argument("--no-open", action="store_true", help="Don't auto-open the Excel file when done.")
    parser.add_argument("-v", "--verbose", action="store_true", help="Debug-level logging.")
    parser.add_argument("-q", "--quiet", action="store_true", help="Warning-level logging only.")
    return parser


def config_from_args(args: argparse.Namespace) -> Config:
    return Config(
        index=args.index,
        start_date=args.start_date,
        end_date=args.end_date,
        years_back=args.years_back,
        time_mode=args.time_mode,
        selected_time=args.selected_time,
        intraday_start=args.intraday_start,
        intraday_end=args.intraday_end,
        intraday_interval_minutes=args.interval_minutes,
        expiry_weekdays=[int(x) for x in args.expiry_weekdays.split(",") if x.strip() != ""],
        request_delay_seconds=args.delay,
        max_retries=args.max_retries,
        output_xlsx=args.output,
        cache_csv=args.cache,
        auto_open=not args.no_open,
    )


def setup_logging(args: argparse.Namespace) -> None:
    level = logging.DEBUG if args.verbose else logging.WARNING if args.quiet else logging.INFO
    logging.basicConfig(level=level, format="%(asctime)s %(levelname)s %(message)s", datefmt="%H:%M:%S")


def run(config: Config, token_box: TokenBox) -> None:
    config.validate()

    start_date, end_date = resolve_date_range(config)
    intraday_times = generate_intraday_times(
        config.intraday_start, config.intraday_end, config.intraday_interval_minutes
    )
    all_tasks = build_task_list(start_date, end_date, config, intraday_times)

    cached_keys = load_cached_keys(config)
    todo = [t for t in all_tasks if t not in cached_keys]

    n_dates = len(set(t[0] for t in all_tasks))
    log.info("Date range: %s to %s (%d weekdays)", start_date, end_date, n_dates)
    log.info("Time mode: %s", config.time_mode)
    log.info(
        "Total snapshots: %d | Already cached: %d | Remaining: %d",
        len(all_tasks), len(cached_keys), len(todo),
    )

    if not todo:
        log.info("Nothing new to fetch -- building Excel from existing cache.")
        build_excel_from_cache(config)
        if config.auto_open:
            auto_open(os.path.abspath(config.output_xlsx))
        return

    session = requests.Session()
    fetched_ok = 0
    fetched_empty = 0

    try:
        for i, (date_str, time_str) in enumerate(todo, start=1):
            status, rows = fetch_snapshot_with_retry(session, token_box, config, date_str, time_str)

            if status == "abort":
                log.warning("Stopped by user. Progress so far is saved in the cache.")
                if fetched_ok or fetched_empty:
                    build_excel_from_cache(config)
                sys.exit(0)

            if status == "ok":
                append_to_cache(config, rows)
                fetched_ok += 1
            else:
                mark_snapshot_done_with_no_data(config, date_str, time_str)
                fetched_empty += 1

            if i % 20 == 0 or i == len(todo):
                log.info("progress: %d/%d (with data: %d, empty/holiday: %d)", i, len(todo), fetched_ok, fetched_empty)

            time.sleep(config.request_delay_seconds)
    except KeyboardInterrupt:
        log.warning("Interrupted -- saving progress fetched so far.")
        if fetched_ok or fetched_empty:
            build_excel_from_cache(config)
        sys.exit(1)

    log.info("Done fetching. With data: %d, empty/holiday: %d", fetched_ok, fetched_empty)
    build_excel_from_cache(config)
    if config.auto_open:
        auto_open(os.path.abspath(config.output_xlsx))


def main(argv: Optional[List[str]] = None) -> None:
    parser = build_arg_parser()
    args = parser.parse_args(argv)
    setup_logging(args)

    config = config_from_args(args)
    try:
        config.validate()
    except ValueError as exc:
        parser.error(str(exc))

    token = resolve_token(args.token)
    token_box = TokenBox(token)
    run(config, token_box)


if __name__ == "__main__":
    main()
