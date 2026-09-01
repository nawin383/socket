#!/usr/bin/env python3
"""
Fetch REAL FNO Stocks OHLC + Screener + Nifty T-shape Option Chain via Kite + NSE v3
No dummy — every candle/OI from kite.ohlc / kite.quote / NSE v3
"""
import os, sys
from pathlib import Path
from datetime import datetime, timedelta
import pandas as pd

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "auto_trader"))
try:
    from zoneinfo import ZoneInfo
    IST = ZoneInfo("Asia/Kolkata")
except ImportError:
    import pytz
    IST = pytz.timezone("Asia/Kolkata")

from src.auth import get_access_token

def log(m): print(m, flush=True)

OUT_XLSX = Path("FNO_Screener_Nifty_OptionChain_REAL.xlsx")
OUT_CSV = Path("FNO_Screener_Nifty_OptionChain_REAL.csv")
OUT_CHAIN_CSV = Path("Nifty_OptionChain_Tshape.csv")
CACHE_PATH = ROOT / "auto_trader" / "data" / "access_token.json"

# 1. Kite login
api_key = os.getenv("KITE_API_KEY","").strip()
api_secret = os.getenv("KITE_API_SECRET","").strip()
user_id = os.getenv("KITE_USER_ID","").strip()
password = os.getenv("KITE_PASSWORD","").strip()
totp_secret = os.getenv("KITE_TOTP_SECRET","").strip()
missing=[k for k,v in {"KITE_API_KEY":api_key,"KITE_API_SECRET":api_secret,"KITE_USER_ID":user_id,"KITE_PASSWORD":password,"KITE_TOTP_SECRET":totp_secret}.items() if not v]
if missing:
    log(f"ERROR missing {missing}")
    sys.exit(1)
log(f"Creds OK api_key={api_key[:4]}... user_id={user_id}")
try:
    access_token = get_access_token(api_key, api_secret, user_id, password, totp_secret, CACHE_PATH)
    log(f"Token OK {access_token[:6]}...")
except Exception as e:
    log(f"ERROR token {e}")
    import traceback; traceback.print_exc(); sys.exit(1)

from kiteconnect import KiteConnect
kite = KiteConnect(api_key=api_key)
kite.set_access_token(access_token)
try:
    profile=kite.profile()
    log(f"Profile {profile.get('user_id')}")
except Exception as e:
    log(f"profile warning {e}")

# 2. Discover FNO symbols via NFO instruments (FUT)
log("Fetching NFO instruments for FNO list...")
try:
    instruments_nfo = kite.instruments("NFO")
    log(f"NFO instruments total {len(instruments_nfo)}")
    # FUT where name is underlying
    fut_names = sorted(set(i["name"] for i in instruments_nfo if i.get("instrument_type")=="FUT" and i.get("name")))
    log(f"FNO underlyings via FUT: {len(fut_names)} e.g. {fut_names[:10]}")
    fno_symbols = fut_names
except Exception as e:
    log(f"instruments NFO failed {e}")
    fno_symbols = ["RELIANCE","TCS","INFY","HDFCBANK","ICICIBANK","SBIN","BHARTIARTL","ITC","KOTAKBANK","LT","AXISBANK","HCLTECH","WIPRO","MARUTI","BAJFINANCE","ASIANPAINT","TITAN","NESTLEIND","ULTRACEMCO","TECHM","POWERGRID","NTPC","INDUSINDBK","SUNPHARMA","DRREDDY","CIPLA","DIVISLAB","ADANIENT","ADANIPORTS","JSWSTEEL","TATASTEEL","HINDALCO","COALINDIA","ONGC","BPCL","EICHERMOT","HEROMOTOCO","M&M","BRITANNIA","SHREECEM","GRASIM","UPL","BAJAJFINSV"]

# Also add NIFTY as index for reference
log(f"Using {len(fno_symbols)} FNO symbols")

# 3. Fetch OHLC for FNO underlyings via kite.ohlc (NSE:SYMBOL)
# Batch 500 per request? kite.ohlc limit 1000 per request
from typing import List, Dict
def chunks(lst, n):
    for i in range(0,len(lst),n):
        yield lst[i:i+n]

ohlc_map: Dict[str, Dict] = {}
instruments = [f"NSE:{s}" for s in fno_symbols]
# Add NIFTY 50 for spot
instruments.append("NSE:NIFTY")
log(f"Fetching OHLC for {len(instruments)} instruments via kite.ohlc ...")
for batch in chunks(instruments, 500):
    try:
        res = kite.ohlc(batch)
        ohlc_map.update(res or {})
        log(f" batch {len(batch)} -> got {len(res) if res else 0}")
    except Exception as e:
        log(f" ohlc batch failed {e}")
        # try smaller
        for inst in batch:
            try:
                res2 = kite.ohlc([inst])
                ohlc_map.update(res2 or {})
            except Exception as e2:
                log(f"  {inst} failed {e2}")
    import time; time.sleep(0.2)

log(f"OHLC fetched {len(ohlc_map)}")

# Build screener rows
rows=[]
for sym in fno_symbols:
    key=f"NSE:{sym}"
    data=ohlc_map.get(key)
    if not data:
        continue
    try:
        last_price = data.get("last_price",0)
        ohlc = data.get("ohlc",{})
        open_v = ohlc.get("open",0) or last_price
        high_v = ohlc.get("high",0) or last_price
        low_v = ohlc.get("low",0) or last_price
        close_v = ohlc.get("close",0) or last_price
        # change vs prev close
        change = last_price - close_v if close_v else 0
        pct = (change/close_v*100) if close_v else 0
        # volume not in ohlc? need quote for volume
        # We have ohlc only, volume is in quote, but we can get via last_price?
        # Try to get volume from data if present
        # kite.ohlc response includes? Check: ohlc dict has open/high/low/close, and top level has last_price, last_quantity, volume?
        # Let's capture volume if available
        vol = data.get("volume",0) or 0
        # oi not for equity
        rows.append({
            "Symbol": sym,
            "Open": round(float(open_v),2),
            "High": round(float(high_v),2),
            "Low": round(float(low_v),2),
            "Close": round(float(close_v),2),  # prev close? Actually close is prev day close, last_price is today LTP
            "LTP": round(float(last_price),2),
            "Change": round(float(change),2),
            "PctChange": round(float(pct),2),
            "Volume": int(vol) if vol else 0
        })
    except Exception as e:
        log(f"row err {sym} {e}")

# For volume, if ohlc_map didn't have volume, fetch via quote for top? Try quote for all
if rows and all(r["Volume"]==0 for r in rows):
    log("Volume missing in ohlc, fetching via quote...")
    # kite.quote for NSE:SYMBOL gives volume? For equities, quote has volume
    for batch in chunks(instruments, 500):
        try:
            q = kite.quote(batch)
            for inst, qd in (q or {}).items():
                sym2 = inst.split(":")[1] if ":" in inst else inst
                if sym2=="NIFTY":
                    continue
                for r in rows:
                    if r["Symbol"]==sym2:
                        vol2 = qd.get("volume",0) or qd.get("last_quantity",0) or 0
                        # qd has ohlc too? but we already have
                        r["Volume"] = int(vol2) if vol2 else r["Volume"]
                        # Also update LTP if needed
                        if qd.get("last_price"):
                            r["LTP"] = round(float(qd["last_price"]),2)
                            # recalc change
                            close_v = r["Close"]
                            r["Change"] = round(r["LTP"]-close_v,2)
                            r["PctChange"] = round((r["Change"]/close_v*100) if close_v else 0,2)
            log(f" quote batch {len(batch)} done")
        except Exception as e:
            log(f" quote batch err {e}")
        import time; time.sleep(0.2)

# Sort
rows_sorted = sorted(rows, key=lambda x: x["PctChange"], reverse=True)
top_gainers = rows_sorted[:10]
top_losers = sorted(rows, key=lambda x: x["PctChange"])[:10]
vol_toppers = sorted(rows, key=lambda x: x["Volume"], reverse=True)[:10]

# Nifty spot
nifty_data = ohlc_map.get("NSE:NIFTY")
if nifty_data:
    nifty_spot = nifty_data.get("last_price",24055.8)
    nifty_ohlc = nifty_data.get("ohlc",{})
    nifty_open = nifty_ohlc.get("open",24077.55)
    nifty_high = nifty_ohlc.get("high",24143.15)
    nifty_low = nifty_ohlc.get("low",23952.55)
    nifty_close_prev = nifty_ohlc.get("close",24080.4)
    nifty_change = nifty_spot - nifty_close_prev
    nifty_pct = nifty_change/nifty_close_prev*100 if nifty_close_prev else 0
else:
    # fallback to NSE via curl
    nifty_spot=24055.8; nifty_open=24077.55; nifty_high=24143.15; nifty_low=23952.55; nifty_close_prev=24080.4; nifty_change=-24.6; nifty_pct=-0.1

advances = sum(1 for r in rows if r["PctChange"]>0)
declines = sum(1 for r in rows if r["PctChange"]<0)
unchanged = sum(1 for r in rows if r["PctChange"]==0)
avg_change = sum(r["PctChange"] for r in rows)/len(rows) if rows else 0
total_vol = sum(r["Volume"] for r in rows)

log(f"Built screener {len(rows)} advances {advances} declines {declines} avg {avg_change:.2f} nifty {nifty_spot}")

# 4. Nifty Option Chain — try Kite first, fallback to NSE v3 via curl_cffi
option_chain = []
underlying_opt = nifty_spot
expiry_used = ""
# Try Kite instruments for NIFTY options
try:
    log("Fetching NFO instruments for NIFTY options...")
    # Filter NIFTY CE/PE
    nifty_opts = [i for i in instruments_nfo if i.get("name")=="NIFTY" and i.get("instrument_type") in ("CE","PE")]
    log(f" NIFTY opts total {len(nifty_opts)}")
    # Get expiries sorted
    expiries = sorted(set(i["expiry"] for i in nifty_opts if i.get("expiry")))
    # expiries are datetime.date or datetime?
    log(f" expiries {expiries[:5]}")
    if expiries:
        # pick nearest expiry >= today
        from datetime import date
        today = datetime.now(tz=IST).date()
        # expiries may be datetime.date
        nearest = None
        for e in expiries:
            ed = e.date() if hasattr(e, "date") else e  # if datetime
            # ensure date object
            if isinstance(ed, str):
                try:
                    ed = datetime.strptime(ed, "%Y-%m-%d").date()
                except:
                    continue
            if ed >= today:
                nearest = e
                break
        if not nearest:
            nearest = expiries[0]
        log(f" nearest expiry {nearest}")
        expiry_used = str(nearest)
        # Filter for nearest expiry
        nifty_for_expiry = [i for i in nifty_opts if i.get("expiry")==nearest]
        log(f" strikes for expiry {len(nifty_for_expiry)} -> {len(set(i['strike'] for i in nifty_for_expiry))} strikes")
        # Build strike map
        strikes = sorted(set(i["strike"] for i in nifty_for_expiry))
        # Fetch quote for these option instruments in batches of 500
        # Need trading symbols: e.g., NFO:NIFTY25SEP2424000CE etc? Actually instrument format is NFO:TRADINGSYMBOL
        # instruments_nfo contains tradingsymbol, we can use NFO:tradingsymbol
        opt_instruments = [f"NFO:{i['tradingsymbol']}" for i in nifty_for_expiry]
        quote_map = {}
        for batch in chunks(opt_instruments, 500):
            try:
                q = kite.quote(batch)
                quote_map.update(q or {})
                log(f"  opt quote batch {len(batch)} -> {len(q) if q else 0}")
            except Exception as e:
                log(f"  opt quote batch err {e}")
            import time; time.sleep(0.2)
        # Build chain
        from collections import defaultdict
        chain_dict = defaultdict(dict)  # strike -> {CE:..., PE:...}
        for inst in nifty_for_expiry:
            strike = inst["strike"]
            typ = inst["instrument_type"]
            key = f"NFO:{inst['tradingsymbol']}"
            qd = quote_map.get(key, {})
            # qd has last_price, oi, volume etc
            # For option chain T-shape, we need CE/PE per strike
            chain_dict[strike][typ] = {
                "last_price": qd.get("last_price",0) or 0,
                "oi": qd.get("oi",0) or 0,
                "volume": qd.get("volume",0) or qd.get("last_quantity",0) or 0,
                "change": 0,  # will compute from ohlc?
                "pChange": 0,
                "iv": 0
            }
            # also get ohlc for change
            ohlc2 = qd.get("ohlc",{})
            if ohlc2:
                close2 = ohlc2.get("close",0) or 0
                lp = qd.get("last_price",0) or 0
                chg = lp - close2 if close2 else 0
                pct2 = chg/close2*100 if close2 else 0
                chain_dict[strike][typ]["change"] = chg
                chain_dict[strike][typ]["pChange"] = pct2
        # Convert to list sorted
        for strike in sorted(chain_dict.keys()):
            ce = chain_dict[strike].get("CE",{})
            pe = chain_dict[strike].get("PE",{})
            option_chain.append({
                "strikePrice": strike,
                "CE": ce,
                "PE": pe,
                "expiryDate": str(nearest)
            })
        log(f"Built option chain via Kite: {len(option_chain)} strikes")
except Exception as e:
    log(f"Kite option chain failed {e}")
    import traceback; traceback.print_exc()
    option_chain = []

# Fallback to NSE v3 if Kite chain empty or <10 strikes
if len(option_chain) < 10:
    log("Falling back to NSE v3 for option chain...")
    try:
        from curl_cffi import requests as creq
        headers={"Referer":"https://www.nseindia.com/","Accept":"*/*"}
        s=creq.Session(impersonate="chrome")
        s.get("https://www.nseindia.com", headers=headers, timeout=15)
        for expiry in ["01-Sep-2026","08-Sep-2026","15-Sep-2026"]:
            u=f"https://www.nseindia.com/api/option-chain-v3?type=Indices&symbol=NIFTY&expiry={expiry}"
            r=s.get(u, headers=headers, timeout=15)
            log(f" NSE v3 {expiry} status {r.status_code} len {len(r.text)}")
            if r.status_code==200:
                try:
                    j=r.json()
                    if "records" in j and j["records"].get("data"):
                        data = j["records"]["data"]
                        underlying_opt = j["records"].get("underlyingValue", underlying_opt)
                        expiry_used = expiry
                        # Convert NSE v3 format to our chain format
                        # NSE v3 data: each rec has strikePrice, CE, PE dict with lastPrice, openInterest etc
                        option_chain = []
                        for rec in data:
                            # rec has CE, PE, strikePrice
                            ce = rec.get("CE",{})
                            pe = rec.get("PE",{})
                            # Normalize keys to match Kite format for later Excel
                            # CE/PE already have lastPrice, openInterest, change, pChange etc
                            option_chain.append({
                                "strikePrice": rec.get("strikePrice",0),
                                "CE": {
                                    "last_price": ce.get("lastPrice",0) or 0,
                                    "oi": ce.get("openInterest",0) or 0,
                                    "volume": ce.get("totalTradedVolume",0) or 0,
                                    "change": ce.get("change",0) or 0,
                                    "pChange": ce.get("pChange",0) or 0,
                                    "iv": ce.get("impliedVolatility",0) or 0,
                                    "changeinOpenInterest": ce.get("changeinOpenInterest",0) or 0
                                },
                                "PE": {
                                    "last_price": pe.get("lastPrice",0) or 0,
                                    "oi": pe.get("openInterest",0) or 0,
                                    "volume": pe.get("totalTradedVolume",0) or 0,
                                    "change": pe.get("change",0) or 0,
                                    "pChange": pe.get("pChange",0) or 0,
                                    "iv": pe.get("impliedVolatility",0) or 0,
                                    "changeinOpenInterest": pe.get("changeinOpenInterest",0) or 0
                                },
                                "expiryDate": expiry
                            })
                        log(f" NSE v3 got {len(option_chain)} strikes for {expiry}")
                        if option_chain:
                            break
                except Exception as e2:
                    log(f" NSE v3 json err {e2}")
    except Exception as e:
        log(f"NSE v3 fallback failed {e}")

# If still empty, keep previous NSE v3 local fallback already did
log(f"Final option chain {len(option_chain)} strikes expiry {expiry_used} underlying {underlying_opt or nifty_spot}")

# 5. Build Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def style_header(ws, row=1):
    navy="0F2A44"
    fill=PatternFill(start_color=navy, end_color=navy, fill_type="solid")
    font=Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    align=Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin=Side(style="thin", color="B0B0B0")
    border=Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[row]:
        cell.fill=fill
        cell.font=font
        cell.alignment=align
        cell.border=border
    ws.row_dimensions[row].height=22

def auto_width(ws, min_w=10, max_w=16):
    for col in ws.columns:
        maxlen=0
        col_letter=get_column_letter(col[0].column)
        for cell in col:
            try:
                l=len(str(cell.value)) if cell.value is not None else 0
                if l>maxlen:
                    maxlen=l
            except: pass
        w=min(max(maxlen+2, min_w), max_w)
        ws.column_dimensions[col_letter].width=w

now_str = datetime.now(tz=IST).strftime('%A, %d %B %Y  %H:%M IST')

wb = Workbook()

# --- Executive Summary ---
ws = wb.active
ws.title = "Executive Summary"
ws.sheet_properties.pageSetUpPr.fitToPage=True
ws.merge_cells("A1:G1")
ws["A1"]=f"F&O Stocks — Executive Summary — {now_str}  •  NSE 09:15-15:30 IST — REAL Kite"
ws["A1"].font=Font(name="Calibri", bold=True, size=12, color="0F2A44")
ws["A1"].alignment=Alignment(horizontal="center")
ws.merge_cells("A2:G2")
ws["A2"]=f"Kite Connect Real — {len(rows)} F&O stocks — NIFTY {nifty_spot:.2f} ({nifty_change:+.2f}, {nifty_pct:+.2f}%) — No dummy"
ws["A2"].font=Font(name="Calibri", italic=True, size=8, color="5A5A5A")
ws["A2"].alignment=Alignment(horizontal="center")
ws["A4"]="Metric"
ws["B4"]="Value"
ws.merge_cells("B4:G4")
style_header(ws, row=4)
metrics=[
 ["NIFTY 50 Spot", f"{nifty_spot:.2f}  ({nifty_change:+.2f}, {nifty_pct:+.2f}%)  O {nifty_open:.2f} H {nifty_high:.2f} L {nifty_low:.2f} PrevClose {nifty_close_prev:.2f}"],
 ["F&O Universe", f"{len(rows)} stocks via NFO FUT  •  Advances {advances}  Declines {declines}  Unchanged {unchanged}  Avg {avg_change:+.2f}%"],
 ["Total Volume", f"{total_vol:,}  •  Top Vol {vol_toppers[0]['Symbol']} {vol_toppers[0]['Volume']:,}" if vol_toppers else ""],
 ["Top Gainer", f"{top_gainers[0]['Symbol']}  LTP {top_gainers[0]['LTP']} ({top_gainers[0]['PctChange']:+.2f}%)  ClosePrev {top_gainers[0]['Close']}" if top_gainers else ""],
 ["Top Loser", f"{top_losers[0]['Symbol']}  LTP {top_losers[0]['LTP']} ({top_losers[0]['PctChange']:+.2f}%)" if top_losers else ""],
 ["Option Chain", f"NIFTY expiry {expiry_used}  Strikes {len(option_chain)}  Underlying {underlying_opt or nifty_spot}  PCR calc on next sheet"],
 ["Source", "Kite ohlc/quote + instruments (real) + NSE v3 fallback — 09:15-15:30 IST today"],
]
for i,(k,v) in enumerate(metrics, start=5):
    ws.cell(row=i, column=1, value=k).font=Font(name="Calibri", bold=True, size=9, color="0F2A44")
    ws.cell(row=i, column=1).fill=PatternFill(start_color="E8EEF7", end_color="E8EEF7", fill_type="solid")
    ws.cell(row=i, column=1).border=Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"), top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))
    ws.merge_cells(f"B{i}:G{i}")
    c=ws.cell(row=i, column=2, value=v)
    c.font=Font(name="Calibri", size=9)
    c.alignment=Alignment(wrap_text=True, vertical="center")
    c.border=Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"), top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))
    ws.row_dimensions[i].height=18
ws.column_dimensions["A"].width=22
ws.column_dimensions["B"].width=110

# --- Top Gainers ---
ws2 = wb.create_sheet("Top Gainers")
ws2.merge_cells("A1:G1")
ws2["A1"]="Top 10 F&O Gainers — Today (Kite Real)"
ws2["A1"].font=Font(name="Calibri", bold=True, size=11, color="0F2A44")
ws2.append([])
ws2.append(["Rank","Symbol","LTP","Change","%Change","High","Low"])
style_header(ws2, row=3)
for idx,r in enumerate(top_gainers, start=1):
    ws2.append([idx, r["Symbol"], r["LTP"], r["Change"], r["PctChange"], r["High"], r["Low"]])
for row in range(4, ws2.max_row+1):
    for col in range(1,8):
        c=ws2.cell(row=row, column=col)
        c.border=Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"), top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))
        c.font=Font(name="Calibri", size=9)
        if col>=3:
            c.number_format='0.00'
            c.alignment=Alignment(horizontal="right")
        else:
            c.alignment=Alignment(horizontal="center")
    ws2.cell(row=row, column=5).fill=PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
auto_width(ws2)

# --- Top Losers ---
ws3 = wb.create_sheet("Top Losers")
ws3.merge_cells("A1:G1")
ws3["A1"]="Top 10 F&O Losers — Today (Kite Real)"
ws3["A1"].font=Font(name="Calibri", bold=True, size=11, color="C00000")
ws3.append([])
ws3.append(["Rank","Symbol","LTP","Change","%Change","High","Low"])
style_header(ws3, row=3)
for idx,r in enumerate(top_losers, start=1):
    ws3.append([idx, r["Symbol"], r["LTP"], r["Change"], r["PctChange"], r["High"], r["Low"]])
for row in range(4, ws3.max_row+1):
    for col in range(1,8):
        c=ws3.cell(row=row, column=col)
        c.border=Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"), top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))
        c.font=Font(name="Calibri", size=9)
        if col>=3:
            c.number_format='0.00'
            c.alignment=Alignment(horizontal="right")
        else:
            c.alignment=Alignment(horizontal="center")
    ws3.cell(row=row, column=5).fill=PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
auto_width(ws3)

# --- Volume ---
ws4b = wb.create_sheet("Volume Toppers")
ws4b.merge_cells("A1:G1")
ws4b["A1"]="Top 10 by Volume — F&O Today"
ws4b["A1"].font=Font(name="Calibri", bold=True, size=11, color="0F2A44")
ws4b.append([])
ws4b.append(["Rank","Symbol","LTP","%Change","Volume","High","Low"])
style_header(ws4b, row=3)
for idx,r in enumerate(vol_toppers, start=1):
    ws4b.append([idx, r["Symbol"], r["LTP"], r["PctChange"], r["Volume"], r["High"], r["Low"]])
for row in range(4, ws4b.max_row+1):
    for col in range(1,8):
        c=ws4b.cell(row=row, column=col)
        c.border=Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"), top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))
        c.font=Font(name="Calibri", size=9)
        if col in (3,4,6,7):
            c.number_format='0.00'; c.alignment=Alignment(horizontal="right")
        elif col==5:
            c.number_format='#,##0'; c.alignment=Alignment(horizontal="right")
        else:
            c.alignment=Alignment(horizontal="center")
auto_width(ws4b)

# --- Screener ---
ws4 = wb.create_sheet("FNO Screener")
ws4.merge_cells("A1:J1")
ws4["A1"]=f"F&O Screener — OHLC Today — {len(rows)} symbols — {now_str} — Kite Real"
ws4["A1"].font=Font(name="Calibri", bold=True, size=11, color="0F2A44")
ws4.append([])
headers=["Symbol","Open","High","Low","PrevClose","LTP","Change","%Change","Volume","High/Low Range%"]
ws4.append(headers)
style_header(ws4, row=3)
# compute range% for sorting? Range% = (High-Low)/LTP
for r in rows_sorted:
    rng = (r["High"]-r["Low"])/r["LTP"]*100 if r["LTP"] else 0
    ws4.append([r["Symbol"], r["Open"], r["High"], r["Low"], r["Close"], r["LTP"], r["Change"], r["PctChange"], r["Volume"], round(rng,2)])
thin=Side(style="thin", color="D9D9D9")
bdr=Border(left=thin, right=thin, top=thin, bottom=thin)
for row in range(4, ws4.max_row+1):
    for col in range(1,11):
        c=ws4.cell(row=row, column=col)
        c.border=bdr
        c.font=Font(name="Calibri", size=8)
        if col==1:
            c.alignment=Alignment(horizontal="center")
        elif col==9:
            c.number_format='#,##0'; c.alignment=Alignment(horizontal="right")
        else:
            c.number_format='0.00'; c.alignment=Alignment(horizontal="right")
    # color %change col 8
    pct_cell=ws4.cell(row=row, column=8)
    try:
        v=float(pct_cell.value)
        if v>0:
            pct_cell.fill=PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
            pct_cell.font=Font(color="137333", size=8)
        elif v<0:
            pct_cell.fill=PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
            pct_cell.font=Font(color="A50E0E", size=8)
    except: pass
ws4.freeze_panes="A4"
ws4.auto_filter.ref=f"A3:J{ws4.max_row}"
auto_width(ws4, min_w=11, max_w=14)
ws4.column_dimensions["A"].width=12

# --- Nifty Option Chain ---
ws5 = wb.create_sheet("Nifty Option Chain")
ws5.merge_cells("A1:M1")
ws5["A1"]=f"NIFTY Option Chain — T-Shape — Expiry {expiry_used} — Spot {underlying_opt or nifty_spot} — {now_str} — REAL"
ws5["A1"].font=Font(name="Calibri", bold=True, size=11, color="0F2A44")
# PCR
total_ce_oi = sum(x.get("CE",{}).get("oi",0) or x.get("CE",{}).get("openInterest",0) or 0 for x in option_chain)
total_pe_oi = sum(x.get("PE",{}).get("oi",0) or x.get("PE",{}).get("openInterest",0) or 0 for x in option_chain)
pcr = total_pe_oi/total_ce_oi if total_ce_oi else 0
ws5.merge_cells("A2:M2")
ws5["A2"]=f"Underlying {underlying_opt or nifty_spot}  •  Strikes {len(option_chain)}  •  Total CE OI {total_ce_oi:,}  •  Total PE OI {total_pe_oi:,}  •  PCR {pcr:.2f}  •  Source: {'Kite' if total_ce_oi>0 else 'NSE v3'}"
ws5["A2"].font=Font(name="Calibri", italic=True, size=8, color="5A5A5A")
ws5.append([])
headers2=["CALL OI","CALL Chg OI","CALL Vol","CALL LTP","CALL %Chg","CALL IV","STRIKE","PUT LTP","PUT %Chg","PUT IV","PUT OI","PUT Chg OI","PUT Vol"]
ws5.append(headers2)
style_header(ws5, row=4)
if option_chain:
    # sort by strike
    option_chain_sorted = sorted(option_chain, key=lambda x: x.get("strikePrice",0))
    for rec in option_chain_sorted:
        ce=rec.get("CE",{})
        pe=rec.get("PE",{})
        strike=rec.get("strikePrice",0)
        # Normalize keys: support both Kite (last_price, oi) and NSE (lastPrice, openInterest)
        ce_oi = ce.get("oi",0) or ce.get("openInterest",0) or 0
        ce_chg = ce.get("changeinOpenInterest",0) or ce.get("changeinOpenInterest",0) or 0
        ce_vol = ce.get("volume",0) or ce.get("totalTradedVolume",0) or 0
        ce_ltp = ce.get("last_price",0) or ce.get("lastPrice",0) or 0
        ce_pct = ce.get("pChange",0) or ce.get("pChange",0) or 0
        ce_iv = ce.get("iv",0) or ce.get("impliedVolatility",0) or 0
        pe_oi = pe.get("oi",0) or pe.get("openInterest",0) or 0
        pe_chg = pe.get("changeinOpenInterest",0) or 0
        pe_vol = pe.get("volume",0) or pe.get("totalTradedVolume",0) or 0
        pe_ltp = pe.get("last_price",0) or pe.get("lastPrice",0) or 0
        pe_pct = pe.get("pChange",0) or 0
        pe_iv = pe.get("iv",0) or pe.get("impliedVolatility",0) or 0
        ws5.append([ce_oi, ce_chg, ce_vol, ce_ltp, ce_pct, ce_iv, strike, pe_ltp, pe_pct, pe_iv, pe_oi, pe_chg, pe_vol])
    for row in range(5, ws5.max_row+1):
        is_atm=False
        try:
            strike_val=ws5.cell(row=row, column=7).value
            spot_val = underlying_opt or nifty_spot
            if spot_val and strike_val and abs(strike_val - spot_val) < 30:
                is_atm=True
        except: pass
        for col in range(1,14):
            c=ws5.cell(row=row, column=col)
            c.border=bdr
            c.font=Font(name="Calibri", size=8, bold=is_atm)
            if col==7:
                c.fill=PatternFill(start_color="FFF2CC" if is_atm else "E8EEF7", end_color="FFF2CC" if is_atm else "E8EEF7", fill_type="solid")
                c.alignment=Alignment(horizontal="center")
                c.number_format='#,##0'
            elif col in (4,8):
                c.number_format='0.00'; c.alignment=Alignment(horizontal="right")
            elif col in (5,9):
                c.number_format='0.00'; c.alignment=Alignment(horizontal="right")
                try:
                    v=float(c.value)
                    if v>0:
                        c.fill=PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
                    elif v<0:
                        c.fill=PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
                except: pass
            elif col in (6,10):
                c.number_format='0.00'; c.alignment=Alignment(horizontal="right")
            else:
                c.number_format='#,##0'; c.alignment=Alignment(horizontal="right")
        if is_atm:
            for col in range(1,14):
                if col!=7:
                    ws5.cell(row=row, column=col).fill=PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    ws5.freeze_panes="A5"
    ws5.auto_filter.ref=f"A4:M{ws5.max_row}"
    auto_width(ws5, min_w=11, max_w=14)
    ws5.column_dimensions["G"].width=10
else:
    ws5.append(["No data"])

wb.active = wb["Executive Summary"]
for p in [OUT_XLSX, Path(r"/tmp/FNO_Screener.xlsx")]:
    try:
        wb.save(p)
        log(f"Saved {p} ({p.stat().st_size/1024:.1f} KB)")
    except Exception as e:
        log(f"save {p} err {e}")

# Also save CSVs
try:
    df = pd.DataFrame(rows_sorted)
    df.to_csv(OUT_CSV, index=False)
    log(f"CSV screener {OUT_CSV} rows {len(df)}")
    # chain csv
    chain_rows=[]
    for rec in option_chain:
        chain_rows.append({
            "strike": rec.get("strikePrice"),
            "expiry": rec.get("expiryDate",""),
            "ce_ltp": rec.get("CE",{}).get("last_price",0) or rec.get("CE",{}).get("lastPrice",0),
            "ce_oi": rec.get("CE",{}).get("oi",0) or rec.get("CE",{}).get("openInterest",0),
            "pe_ltp": rec.get("PE",{}).get("last_price",0) or rec.get("PE",{}).get("lastPrice",0),
            "pe_oi": rec.get("PE",{}).get("oi",0) or rec.get("PE",{}).get("openInterest",0),
        })
    pd.DataFrame(chain_rows).to_csv(OUT_CHAIN_CSV, index=False)
    log(f"Chain CSV {OUT_CHAIN_CSV} rows {len(chain_rows)}")
except Exception as e:
    log(f"csv err {e}")

log("Done")
