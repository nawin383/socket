#!/usr/bin/env python3
"""
Fetch REAL India VIX 15min OHLC for last 6 months via Kite Connect historical API.
Uses same auto-login as auto_trader/src/auth.py (TOTP + request_token + access_token).
No dummy data — every candle comes from kite.historical_data.
"""
import os, sys
from pathlib import Path
from datetime import datetime, timedelta
import pandas as pd

# Ensure auto_trader/src on path for auth helper
ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "auto_trader"))

try:
    from zoneinfo import ZoneInfo
    IST = ZoneInfo("Asia/Kolkata")
except ImportError:
    import pytz
    IST = pytz.timezone("Asia/Kolkata")

from src.auth import get_access_token

# Config
OUT_XLSX = Path("India_VIX_15min_6M_REAL.xlsx")
OUT_CSV = Path("India_VIX_15min_6M_REAL.csv")
CACHE_PATH = ROOT / "auto_trader" / "data" / "access_token.json"

def log(m):
    print(m, flush=True)

# 1. Get creds from env (from GitHub Secrets)
api_key = os.getenv("KITE_API_KEY", "").strip()
api_secret = os.getenv("KITE_API_SECRET", "").strip()
user_id = os.getenv("KITE_USER_ID", "").strip()
password = os.getenv("KITE_PASSWORD", "").strip()
totp_secret = os.getenv("KITE_TOTP_SECRET", "").strip()

missing = [k for k,v in {"KITE_API_KEY":api_key,"KITE_API_SECRET":api_secret,"KITE_USER_ID":user_id,"KITE_PASSWORD":password,"KITE_TOTP_SECRET":totp_secret}.items() if not v]
if missing:
    log(f"ERROR missing creds: {missing}")
    sys.exit(1)

log(f"Creds OK: api_key={api_key[:4]}... user_id={user_id}")

# 2. Get fresh access_token via TOTP flow (cached per day)
try:
    access_token = get_access_token(api_key, api_secret, user_id, password, totp_secret, CACHE_PATH)
    log(f"Access token OK: {access_token[:6]}... (cached file {CACHE_PATH})")
except Exception as e:
    log(f"ERROR get_access_token failed: {e}")
    import traceback; traceback.print_exc()
    sys.exit(1)

# 3. Kite client
from kiteconnect import KiteConnect
kite = KiteConnect(api_key=api_key)
kite.set_access_token(access_token)

# Verify
try:
    profile = kite.profile()
    log(f"Kite profile: user_id={profile.get('user_id')} exchanges={profile.get('exchanges')}")
except Exception as e:
    log(f"WARNING profile failed: {e}")

# 4. Discover INDIA VIX instrument_token dynamically
log("Fetching NSE instruments to find INDIA VIX token...")
try:
    instruments = kite.instruments("NSE")
    # Look for INDIA VIX
    vix_candidates = [i for i in instruments if "INDIA VIX" in i.get("tradingsymbol","") or "INDIA VIX" in i.get("name","")]
    log(f"Found {len(vix_candidates)} INDIA VIX candidates")
    for c in vix_candidates[:5]:
        log(str(c))
    if not vix_candidates:
        # fallback: try BSE?
        instruments_bse = kite.instruments("BSE")
        vix_candidates = [i for i in instruments_bse if "VIX" in i.get("tradingsymbol","")]
        log(f"BSE candidates {len(vix_candidates)}")
    if not vix_candidates:
        log("ERROR no INDIA VIX instrument found, trying known token 264512 (common)")
        vix_token = 264512
        vix_symbol = "NSE:INDIA VIX"
    else:
        # pick the one with instrument_type maybe?
        # Prefer tradingsymbol exactly "INDIA VIX"
        exact = [i for i in vix_candidates if i.get("tradingsymbol")=="INDIA VIX"]
        chosen = exact[0] if exact else vix_candidates[0]
        vix_token = chosen["instrument_token"]
        vix_symbol = f"{chosen['exchange']}:{chosen['tradingsymbol']}"
        log(f"Chosen VIX: {chosen}")
except Exception as e:
    log(f"ERROR instruments fetch failed: {e}")
    import traceback; traceback.print_exc()
    sys.exit(1)

log(f"Using VIX token={vix_token} symbol={vix_symbol}")

# 5. Fetch historical 15minute for last 6 months
# Kite historical: from_date and to_date are datetime objects, interval = "15minute"
# Need to handle pagination: Kite allows max ~1000 candles per call? Actually it returns all within range.
# For 6 months ~126 trading days * 25 candles per day (09:15-15:30 = 6h15m = 25 *15m) = ~3150 candles, within one call.
# To be safe, we fetch in 60-day chunks (since Kite may limit to 100 days per call for intraday?)
# We'll do 3 chunks: last 60d, 60-120d, 120-180d

from datetime import datetime

now_ist = datetime.now(tz=IST)
# Kite expects naive or UTC? Use IST naive? We'll pass IST datetime
to_dt = now_ist
from_dt = to_dt - timedelta(days=180)
log(f"Range: {from_dt.date()} -> {to_dt.date()} IST")

# Try single call first
all_candles = []
try:
    # Kite historical_data expects datetime in local? We'll pass as datetime with IST
    # If fails, try chunked
    candles = kite.historical_data(vix_token, from_dt, to_dt, "15minute")
    log(f"Single call returned {len(candles)} candles")
    all_candles = candles
except Exception as e:
    log(f"Single call failed: {e} -> trying chunked 60d")
    all_candles = []
    chunk_days = 60
    cur = from_dt
    while cur < to_dt:
        nxt = min(cur + timedelta(days=chunk_days), to_dt)
        log(f" Chunk {cur.date()} -> {nxt.date()}...")
        try:
            chunk = kite.historical_data(vix_token, cur, nxt, "15minute")
            log(f"  -> {len(chunk)} candles")
            all_candles.extend(chunk)
        except Exception as e2:
            log(f"  chunk failed: {e2}")
        cur = nxt
        import time; time.sleep(0.35)  # pace

# Deduplicate by date
if not all_candles:
    log("ERROR no candles fetched")
    sys.exit(1)

# Sort and dedupe
# candles are dicts with keys: date, open, high, low, close, volume, oi?
# date is datetime
all_candles.sort(key=lambda x: x["date"])
# dedupe by date
seen = set()
uniq = []
for c in all_candles:
    d = c["date"]
    # normalize to string
    key = d.isoformat() if hasattr(d, "isoformat") else str(d)
    if key not in seen:
        seen.add(key)
        uniq.append(c)
all_candles = uniq
log(f"Total unique candles: {len(all_candles)} after dedupe")

df = pd.DataFrame(all_candles)
# Ensure date is datetime with IST
df["date"] = pd.to_datetime(df["date"])
# Convert to IST if needed
try:
    if df["date"].dt.tz is None:
        df["date"] = df["date"].dt.tz_localize("UTC").dt.tz_convert(IST)
    else:
        df["date"] = df["date"].dt.tz_convert(IST)
except Exception as e:
    log(f"tz convert warning: {e}")

df = df.sort_values("date")
log(df.head(3).to_string())
log(df.tail(3).to_string())

# 6. Save to Excel with formatting
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "15m_6M_REAL"
    
    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = f"India VIX (NSE:INDIA VIX) — 15-MINUTE OHLC — REAL KITE DATA — {len(df)} candles — {from_dt.strftime('%d-%b-%Y')} → {to_dt.strftime('%d-%b-%Y')} IST — Token {vix_token}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=11, color="0F2A44")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 18
    
    ws.merge_cells("A2:G2")
    ws["A2"] = f"Source: Kite Connect historical_data(token={vix_token}, interval=15minute) via GitHub Secrets (api_key={api_key[:4]}...) — Generated {datetime.now(tz=IST).strftime('%d-%b-%Y %H:%M IST')} — NO DUMMY DATA"
    ws["A2"].font = Font(name="Calibri", italic=True, size=8, color="5A5A5A")
    
    headers = ["Date (IST)", "Time (IST)", "Open", "High", "Low", "Close", "Volume"]
    ws.append([]) # row3 blank
    ws.append(headers) # row4
    # style header row 4
    navy = "0F2A44"
    fill = PatternFill(start_color=navy, end_color=navy, fill_type="solid")
    font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[4]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border
    ws.row_dimensions[4].height = 22
    
    # Data rows starting row5
    thin2 = Side(style="thin", color="D9D9D9")
    border2 = Border(left=thin2, right=thin2, top=thin2, bottom=thin2)
    for _, row in df.iterrows():
        ts = row["date"]
        date_str = ts.strftime("%d-%m-%Y")
        time_str = ts.strftime("%H:%M:%S")
        ws.append([date_str, time_str, round(float(row["open"]),2), round(float(row["high"]),2), round(float(row["low"]),2), round(float(row["close"]),2), int(row.get("volume",0))])
    # format
    for r in range(5, ws.max_row+1):
        for c in range(1,8):
            cell = ws.cell(row=r, column=c)
            cell.border = border2
            cell.font = Font(name="Calibri", size=9)
            if c in (1,2):
                cell.alignment = Alignment(horizontal="center")
            elif c in (3,4,5,6):
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
        if r % 2 == 0:
            for c in range(1,8):
                ws.cell(row=r, column=c).fill = PatternFill(start_color="F2F6FC", end_color="F2F6FC", fill_type="solid")
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:G{ws.max_row}"
    from openpyxl.utils import get_column_letter
    for col_idx, w in enumerate([14,12,10,10,10,10,12], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    wb.save(OUT_XLSX)
    log(f"Saved Excel {OUT_XLSX} ({OUT_XLSX.stat().st_size/1024:.1f} KB)")
    
    # Also save CSV
    df.to_csv(OUT_CSV, index=False)
    log(f"Saved CSV {OUT_CSV}")
    
    # Log stats
    log(f"Rows: {len(df)} From {df['date'].min()} To {df['date'].max()}")
    log(f"Excel ready for download as artifact")

except Exception as e:
    log(f"ERROR Excel save failed: {e}")
    import traceback; traceback.print_exc()
    sys.exit(1)
